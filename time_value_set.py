import pandas as pd
import openpyxl as pyxl
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import datetime
import time
import os
import shutil
from operator import itemgetter
#import jpholiday

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options


#ーーーーーーーー前回ダウンロードファイル削除ーーーーーーーーーー

dr_files = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/TimeZoneData'
dr_read = os.listdir(dr_files)

print(dr_read)

for file_name in dr_read:
  del_f_path = dr_files + '/' + file_name#削除ファイルパスの設定
  os.remove(del_f_path)#dataf内のファイルの削除
  
  
  
print("日付指定値を入力して下さい")

print("0 = 本日日付" + "\n" + "1 = 日付指定")
swith = input()

if swith == '0' :

  todaytime = datetime.date.today()
  tod = '{0:20%y%m%d}'.format(todaytime)#今日の日付(西暦)

  today_1 = datetime.date.today()
  f_day = today_1.strftime("%Y%m%d")#西暦表記
  y_day = today_1.strftime("%Y")#西暦表記
  m_day = today_1.strftime("%m")#月
  d_day = today_1.strftime("%d")#日
  w_day = today_1.weekday()

        

  print(f_day)
  print(y_day)
  print(m_day)
  print(d_day)
  print(w_day)


elif swith == '1':
  
  print("年度を指定して下さい")
  print("(例) 20XX")
  y_ =input()
  y_day = y_#.strftime("%Y")
  
  print("月を指定して下さい")
  print("(例) 01 ⇒ 1月")
  m_ = input()
  m_day = m_
  
  print("日を指定して下さい")
  print("(例) 01 ⇒ 1日")
  d_ = input()
  d_day = d_

  f_day = str(y_day) + str(m_day) + str(d_day)#.strftime("%Y%m%d")#西暦表記
  print(f_day)
  w_day = datetime.datetime.strptime(f_day,'%Y%m%d').weekday()
    

  print(f_day)
  print(y_day)
  print(m_day)
  print(d_day)
  print(w_day)  
  
else:
  print("既定値エラー")  


week = ['月','火','水','木','金','土','日']

day_type_list = {
  "0" : "平日",
  "1" : "土日祝"
}


kasiwa = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[3]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[3]','柏','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[3]',"01001008 FUN柏","01001008",]
chiba = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[4]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[4]', '千葉','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[4]',"01001009 FUN千葉C-one","01001009",]
isesaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[9]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[9]','伊勢崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[9]',"01001028 FUNスマーク伊勢崎","01001028",]
nagamachi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[11]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[11]','長町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[11]',"01001032 FUNララガーデン長町","01001032",]
hunabashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[12]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[12]','船橋','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[12]',"01001033 FUNららぽーとTOKYO-BAY","01001033",]
hujimi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[13]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[13]','富士見','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[13]',"01001034 FUNららぽーと富士見","01001034",]
reiku = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[15]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[15]','レイク','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[15]',"01001036 FUNイオンレイクタウン","01001036",]
ebina = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[17]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[17]','海老名','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[17]',"01001038 FUNららぽーと海老名","01001038",]
musashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[18]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[18]','むさし','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[18]',"01001039 FUNイオンモールむさし村山","01001039",]
hiratuka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[19]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[19]','平塚','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[19]',"01001040 FUNららぽーと湘南平塚","01001040",]
natori = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[20]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[20]','名取','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[20]',"01001041 FUNイオンモール名取","01001041",]
otaka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[21]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[21]','大高','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[21]',"01001042 FUNイオンモール大高","01001042",]
togocyo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[22]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[22]','東郷町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[22]',"01001043 FUNららぽーと愛知東郷","01001043",]
ota = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[23]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[23]','太田','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[23]',"01001044 FUNイオンモール太田","01001044",]
mito = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[24]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[24]','水戸','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[24]',"01001045 FUNイオンモール水戸内原","01001045",]
expo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[25]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[25]','EXPO','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[25]',"01001046 FUNららぽーとEXPOCITY","01001046",]
kawasaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[26]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[26]','川崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[26]',"01001047 FUNラゾーナ川崎プラザ","01001047",]
sinmisato = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[27]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[27]','新三郷','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[27]',"01001048 FUNららぽーと新三郷","01001048",]
makuhari = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[28]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[28]','幕張','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[28]',"01001049 FUNイオンモール幕張新都心","01001049",]
kagamihara = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[29]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[29]','各務原','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[29]',"01001050 FUNイオンモール各務原","01001050",]
sakai = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[30]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[30]','堺','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[30]',"01001051 FUNららぽーと堺","01001051",]



tenpo_list = [
  kasiwa,
  chiba,
  isesaki,
  # nagamachi,
  # hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara,
  sakai,
  ]


# tenpo_options = {

# 1008:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[3]',#柏
# 1009:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[4]',#千葉
# 1028:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[9]',#伊勢崎
# 1032:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[11]',#長町
# 1033:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[12]',#船橋
# 1034:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[13]',#富士見
# 1036:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[15]',#レイク
# 1038:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[17]',#海老名
# 1039:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[18]',#むさし
# 1040:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[19]',#平塚
# 1041:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[20]',#名取
# 1042:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[21]',#大高
# 1043:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[22]',#東郷町
# 1044:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[23]',#太田
# 1045:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[24]',#水戸
# 1046:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[25]',#EXPO
# 1047:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[26]',#川崎
# 1048:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[27]',#新三郷
# 1049:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[28]',#幕張
# 1050:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[29]',#各務原
# 1051:'//*[@id="ContentPlaceHolder1_DropDownList2"]/option[30]',#各務原

# }

tenpo_options_name = {
  
1008:'柏',
1009:'千葉',
1028:'伊勢崎',
1032:'長町',
1033:'船橋',
1034:'富士見',
1036:'レイク',
1038:'海老名',
1039:'むさし',
1040:'平塚',
1041:'名取',
1042:'大高',
1043:'東郷町',
1044:'太田',
1045:'水戸',
1046:'EXPO',
1047:'川崎',
1048:'新三郷',
1049:'幕張',
1050:'各務原',
1051:'堺',

}

shop_id =[
                                                                                                                                                                                                                                
1008,#柏
1009,#千葉
1028,#伊勢崎
1032,#長町
1033,#船橋
1034,#富士見
1036,#レイク
1038,#海老名
1039,#むさし
1040,#平塚
1041,#名取
1042,#大高
1043,#東郷町
1044,#太田
1045,#水戸
1046,#EXPO
1047,#川崎
1048,#新三郷
1049,#幕張
1050,#各務原
1051,#堺
]


#一致店舗の列調整
tenpo_pitch = {
  '柏':0,
  '千葉':1,
  '伊勢崎':2,
  '長町':3,
  '船橋':4,
  '富士見':5,
  'レイク':6,
  '海老名':7,
  'むさし':8,
  '平塚':9,
  '名取':10,
  '大高':11,
  '東郷町':12,
  '太田':13,
  '水戸':14,
  'エキスポ':15,
  '川崎':16,
  '新三郷':17,
  '幕張':18,
  '各務原':19,
  '堺':20,
}

#def scr():#スクレイピングメソッド
options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')

url = 'http://tri.hanbai-net.com/system/Login.aspx'
#driver = webdriver.Chrome("C:/Users/fun-f/chromedriver.exe")#旧
driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
#driver = webdriver.Chrome('C:/Users/fun-f/Desktop/myfile/chromedriver.exe')

driver.get(url)

id_1 = 'trinityadmin'
id_2 = 'AdminTrinity'

loginid_1 = driver.find_element(By.ID,"ContentPlaceHolder1_txtUserCode")
loginid_2 = driver.find_element(By.ID,"ContentPlaceHolder1_txtPassword")

loginid_1.send_keys(id_1)#ユーザーIDを入力
loginid_2.send_keys(id_2)#パスワードを入力



driver.find_element(By.ID,"ContentPlaceHolder1_btnLogin").click() 
#ログインボタンをクリック
#ログインボタンをクリック

driver.get('http://tri.hanbai-net.com/system/00000000.aspx')

driver.get('http://tri.hanbai-net.com/system/30021101.aspx?id=010199')#時間帯売上



driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").clear()
driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(f_day)

driver.find_element(By.ID,"ContentPlaceHolder1_txtCond01").clear()
driver.find_element(By.ID,"ContentPlaceHolder1_txtCond01").send_keys(f_day)

#driver.find_element(By.ID,"ContentPlaceHolder1_DropDownList2").click()

for id in tenpo_list :
  
  Select_element = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownList2")
  Select_element_set = Select(Select_element)
  Select_element_set.select_by_value(str(id[5]))

  driver.find_element(By.ID,"ContentPlaceHolder1_btnCSV").click()#CSV出力

  time.sleep(5)
  
  filelists = []
  for file in os.listdir("C:/Users/古内翔平/Downloads"):#ディレクトリ内をfor文で取り出す
    base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
    
    if ext == '.csv':#拡張子csvが一致した場合…
      if base == '時間帯売上一覧':
        filelists.append([file, os.path.getctime("C:/Users/古内翔平/Downloads/" + str(file))])#filelistsに取り出したfileにダウンロード時間を追加
        #print("file:{},csv:{}" .format(file,csv))
        filelists.sort(key=itemgetter(0), reverse=True)#
        
        MAX_CNT = 0
        
        for i, file in enumerate(filelists):
          if i > MAX_CNT-1:
            print(file[0])
            #file_1 = os.rename(i[0], 'kasi.csv')
            os.rename("C:/Users/古内翔平/Downloads/" + str(file[0]), "C:/Users/古内翔平/Downloads/" + str(f_day) + str(id[5]) + str(w_day) + '時間帯売上.csv')
            shutil.move("C:/Users/古内翔平/Downloads/" + str(f_day) + str(id[5]) + str(w_day) + '時間帯売上.csv','C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/TimeZoneData')
            
                      
print("スクレイピング完了")   
driver.close()    
  

  
out_file2 = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/シフト管理/予実管理.xlsx"#予実管理出力ファイル

wb_out_file2 = pyxl.load_workbook(out_file2)

ws_out_file2 = wb_out_file2["時間帯実績"]  

for id_n,shop_key2 in zip(tenpo_list,tenpo_pitch):
  # ファイル構成⇒　"2022" + "01" + "01" + "1040" + "5" + "時間帯売上.csv"

  select_file = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/TimeZoneData/' + str(y_day) + str(m_day) + str(d_day) + str(id_n[5]) + str(w_day) + "時間帯売上.csv"
  print(select_file)
  
  r_select_file = pd.read_csv(select_file,encoding="cp932")
  
  print(r_select_file)
  
  time_value_list = r_select_file["販売価格合計"].values
  print(time_value_list)
  
  t_7 = time_value_list[7]
  t_8 = time_value_list[8]
  t_9 = time_value_list[9]
  t_10 = time_value_list[10]
  t_11 = time_value_list[11]
  t_12 = time_value_list[12]
  t_13 = time_value_list[13]
  t_14 = time_value_list[14]
  t_15 = time_value_list[15]
  t_16 = time_value_list[16]
  t_17 = time_value_list[17]
  t_18 = time_value_list[18]
  t_19 = time_value_list[19]
  t_20 = time_value_list[20]
  t_21 = time_value_list[21]
  t_22 = time_value_list[22]
  t_23 = time_value_list[23]
  
  # for shop_key2 in tenpo_pitch:
    # select_data = shop_shift_list[shop_key2]
  for d_no in range(0,31):
    
    select_day = str(y_day) + "-" + str(m_day) +"-" + str(d_day) + " 00:00:00"
    
    print(select_day)
    print("ここ" + str(select_day))
    for i in range(0,366):#364
      target_cell = ws_out_file2["A" + str(369 + i)].value#4⇒369
      
      if str(target_cell) != str(select_day):
        
        print("no")
      
      else:  
        
        cell_r = 369 + i#4⇒369
        pitch = 18 #列間隔
        out_file2_header = cell_r#行番号
        out_file2_col = 4 + (tenpo_pitch[shop_key2] * pitch)#列番号

        ws_out_file2.cell(out_file2_header,out_file2_col).value = t_7
        ws_out_file2.cell(out_file2_header,out_file2_col + 1).value = t_8
        ws_out_file2.cell(out_file2_header,out_file2_col + 2).value = t_9
        ws_out_file2.cell(out_file2_header,out_file2_col + 3).value = t_10
        ws_out_file2.cell(out_file2_header,out_file2_col + 4).value = t_11
        ws_out_file2.cell(out_file2_header,out_file2_col + 5).value = t_12
        ws_out_file2.cell(out_file2_header,out_file2_col + 6).value = t_13
        ws_out_file2.cell(out_file2_header,out_file2_col + 7).value = t_14
        ws_out_file2.cell(out_file2_header,out_file2_col + 8).value = t_15
        ws_out_file2.cell(out_file2_header,out_file2_col + 9).value = t_16
        ws_out_file2.cell(out_file2_header,out_file2_col + 10).value = t_17
        ws_out_file2.cell(out_file2_header,out_file2_col + 11).value = t_18
        ws_out_file2.cell(out_file2_header,out_file2_col + 12).value = t_19
        ws_out_file2.cell(out_file2_header,out_file2_col + 13).value = t_20
        ws_out_file2.cell(out_file2_header,out_file2_col + 14).value = t_21
        ws_out_file2.cell(out_file2_header,out_file2_col + 15).value = t_22
        ws_out_file2.cell(out_file2_header,out_file2_col + 16).value = t_23

  wb_out_file2.save(out_file2)  
