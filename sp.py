import pandas as pd
import openpyxl as xlpy
import os

import selenium
from selenium import webdriver
from operator import itemgetter
import shutil
import datetime
import time
from datetime import timedelta

#ーーーーーーー販売NETスクレイピングーーーーーーーーーーー


kasiwa = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[3]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[3]','柏','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[3]',"01001008"]
chiba = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[4]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[4]', '千葉','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[4]',"01001009"]
isesaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[9]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[9]','伊勢崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[9]',"01001028"]
nagamachi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[11]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[11]','長町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[11]',"01001032"]
hunabashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[12]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[12]','船橋','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[12]',"01001033"]
hujimi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[13]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[13]','富士見','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[13]',"01001034"]
reiku = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[15]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[15]','レイク','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[15]',"01001036"]
ebina = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[17]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[17]','海老名','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[17]',"01001038"]
musashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[18]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[18]','むさし','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[18]',"01001039"]
hiratuka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[19]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[19]','平塚','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[19]',"01001040"]
natori = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[20]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[20]','名取','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[20]',"01001041"]
otaka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[21]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[21]','大高','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[21]',"01001042"]
togocyo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[22]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[22]','東郷町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[22]',"01001043"]
ota = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[23]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[23]','太田','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[23]',"01001044"]
mito = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[24]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[24]','水戸','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[24]',"01001045"]
expo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[25]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[25]','EXPO','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[25]',"01001046"]
kawasaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[26]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[26]','川崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[26]',"01001047"]
sinmisato = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[27]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[27]','新三郷','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[27]',"01001048"]
makuhari = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[28]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[28]','幕張','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[28]',"01001049"]
kagamihara = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[29]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[29]','各務原','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[29]',"01001050"]
sakai = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[30]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[30]','堺','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[30]',"01001051"]


tenpo_list = [
  kasiwa,
  chiba,
  isesaki,
  nagamachi,
  hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara,
  sakai,
  ]
  
tenpo_list_2 = [
  kasiwa,
  chiba,
  isesaki,
  nagamachi,
  hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara,
  sakai,
  ]
  
weekday_writer_list = {
  0:"G",#月曜日
  1:"H",#火曜日
  2:"I",#水曜日
  3:"J",#木曜日
  4:"K",#金曜日
  5:"L",#土曜日
  6:"M" #日曜日
}

del_1 = 'ｼｮｯﾋﾟﾝｸﾞﾊﾞｯｸﾞS'#除外品番
del_2 = "ｼｮｯﾋﾟﾝｸﾞﾊﾞｯｸﾞM"#除外品番
del_3 = "ｼｮｯﾋﾟﾝｸﾞﾊﾞｯｸﾞL"#除外品番
del_4 = "ｼｮｯﾋﾟﾝｸﾞﾊﾞｯｸﾞLL"#除外品番
del_5 = "ﾊﾝﾄﾞｸﾘｰﾝｼﾞｪﾙ"#除外品番
del_6 = "ｷﾚｲﾏｽｸ"#除外品番
del_7 = "ｼｮｯﾋﾟﾝｸﾞﾊﾞｯｸﾞLL＊"#除外品番
del_8 = "わけあり"#除外品番
del_9 = "ｷﾞﾌﾄS"
del_10 = "ｷﾞﾌﾄM"
del_11 = "ｷﾞﾌﾄL"




shopper_folder = "C:/Users/fun-f/Desktop/analysis/shopper"
inventory_folder = "C:/Users/fun-f/Desktop/analysis/inventory"



def download():
  
  del_list = os.listdir(shopper_folder)
  
  for del_file in del_list:
    os.remove(os.path.join(shopper_folder,del_file))
  
  del_list2 = os.listdir(inventory_folder)
  
  for del_file2 in del_list2:
    os.remove(os.path.join(inventory_folder,del_file2))
    
  print("期間のSTART日を入力して下さい！")
  print("例 20XX0101(20XX年1月1日の場合)")
  #period1 = tod + str(input())
  
  todaytime = datetime.date.today()
  period1 = '{0:20%y%m%d}'.format(todaytime)#今日の日付(西暦)

  #period1 = str(input())
  
  # print("期間のEND日を入力して下さい！")
  # print("例 20XX0107(20XX年1月7日の場合)")
  # = tod + str(input())
  #period2 = str(input())


  y = period1[0:4]
  m = period1[4:6]
  d = period1[6:8]

  ymd = datetime.datetime.strptime(str(y) + '-' + str(m) + '-' + str(d), '%Y-%m-%d')#period1を日付に変換
  
  ymd2 = ymd - timedelta(days= 60)
  #1.5ヵ月分の在庫
  
  print(ymd2)
  
  y2 = str(ymd2.year).zfill(4)
  m2 = str(ymd2.month).zfill(2)
  d2 = str(ymd2.day).zfill(2)
  
  period2 = y2 + m2 + d2
  
  
  
  
  
  #point_week = date(int(y),int(m),int(d)).isocalendar().week
  #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
  
  #品番別の売上をダウンロードするプログラム
  
  #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
  
  
  url = 'http://tri.hanbai-net.com/system/Login.aspx'
  driver = webdriver.Chrome("C:/Users/fun-f/chromedriver.exe")#2021 0724

  id_1 = 'trinityadmin'
  id_2 = 'AdminTrinity'

  driver.get(url)

  loginid_1 = driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtUserCode"]')
  loginid_2 = driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtPassword"]')

  loginid_1.send_keys(id_1)#ユーザーIDを入力
  loginid_2.send_keys(id_2)#パスワードを入力


  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnLogin"]').click() 
  #ログインボタンをクリック

  driver.get('http://tri.hanbai-net.com/system/00000000.aspx')


  #品番別売上集計
  driver.get("http://tri.hanbai-net.com/system/30021901.aspx?id=010199")


  #--------店別---------
  #★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
  #品番別売上集計をダウンロード

  for i_1 in tenpo_list:
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownListCond04"]').click()#店舗名指定上段

    driver.find_element_by_xpath(str(i_1[0])).click()#店舗選択


    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownListCond05"]').click()#店舗名指定下段

    driver.find_element_by_xpath(str(i_1[1])).click()#店舗選択
    
    #期間指定
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02"]').clear()
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02"]').send_keys(period2)#開始
    
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond03"]').clear()
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond03"]').send_keys(period1)#終了
  

    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCondRun"]').click()#検索

    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()#CSV出力

    time.sleep(5)#一時待機


    filelists = []
    for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
        base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
        if ext == '.csv':#拡張子csvが一致した場合…
            if base == '品番売上集計':
                filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
                #print("file:{},csv:{}" .format(file,csv))
                filelists.sort(key=itemgetter(0), reverse=True)#
                MAX_CNT = 0
                for i, file in enumerate(filelists):
                    if i > MAX_CNT-1:
                        print(file[0])
                        #file_1 = os.rename(i[0], 'kasi.csv')
                        os.rename(file[0], str(i_1[2]) + '.csv')
                        shutil.move(str(i_1[2]) + '.csv',shopper_folder) 
                        
    #在庫一覧表
  driver.get("http://tri.hanbai-net.com/system/21026001.aspx?id=010199")


  #--------店別---------
  #★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
  #品番別売上集計をダウンロード

  #日付入力
  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond04"]').clear()    
  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond04"]').send_keys(period1)
  
  #CSV出力
  # driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()
  
  # time.sleep(5)#一時待機


  # filelists = []
  # for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
  #     base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
  #     if ext == '.csv':#拡張子csvが一致した場合…
  #         if base == '在庫一覧_':
  #             filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
  #             #print("file:{},csv:{}" .format(file,csv))
  #             filelists.sort(key=itemgetter(0), reverse=True)#
  #             MAX_CNT = 0
  #             for i, file in enumerate(filelists):
  #                 if i > MAX_CNT-1:
  #                     print(file[0])
  #                     #file_1 = os.rename(i[0], 'kasi.csv')
  #                     os.rename(file[0],'全店.csv')
  #                     shutil.move('全店.csv',inventory_folder) 
                      
  #                     time.sleep(7)
  #driver.find_element_xpath("")
  for i_1 in tenpo_list:
    #店舗入力
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownListCond01"]').send_keys(i_1[4])
    
    #日付入力
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond04"]').clear()    
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond04"]').send_keys(period1)
    
    #CSV出力
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()
    
    time.sleep(5)#一時待機


    filelists = []
    for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
        base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
        if ext == '.csv':#拡張子csvが一致した場合…
            if base == '在庫一覧_':
                filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
                #print("file:{},csv:{}" .format(file,csv))
                filelists.sort(key=itemgetter(0), reverse=True)#
                MAX_CNT = 0
                for i, file in enumerate(filelists):
                    if i > MAX_CNT-1:
                        print(file[0])
                        #file_1 = os.rename(i[0], 'kasi.csv')
                        os.rename(file[0], str(i_1[2]) + '.csv')
                        shutil.move(str(i_1[2]) + '.csv',inventory_folder)                     

  driver.close()
    
  
def totalling():
  
  import pandas as pd
  import openpyxl as pyxl
  import os
  index_no = 4
  class shopper_totalling:
    
    def __init__(self,shopname,filename, count,inventory,items):
      
      self.shopname = shopname
      self.filename = filename
      self.count = count
      self.inventory = inventory
      self.items:list = items
 
  
  file_list = os.listdir(shopper_folder)
    
    
  #index_no = 2
  for shop_name in tenpo_list :
    
    for file_name in file_list:
      if shop_name[2] in file_name :
        print(file_name)
        
        r_file = pd.read_csv(os.path.join(shopper_folder,file_name),encoding="SHIFT-JIS")
        
        df_r_file = pd.DataFrame(r_file)
        
        
        item_cd = pd.DataFrame(df_r_file["商品コード"].astype(str).str.zfill(10).values,columns=["商品CD"])  
        item_name = pd.DataFrame(df_r_file["商品名"].values,columns=["商品名"])
        category_cd = pd.DataFrame(df_r_file["商品コード"].astype(str).str.zfill(10).str[2:4].values,columns=["アイテムCD"])
        quantity = pd.DataFrame(df_r_file["合計数量"].values,columns=["数量"])
        
        item_list = pd.concat([item_cd,item_name,category_cd,quantity],axis=1)
        
        shopper_list = item_list[item_list["アイテムCD"].values == "98"]
        
        #Sサイズ
        try:
          shopper_S = item_list[item_list["商品CD"].values == "9998998006"]
          shopper_S_quantity = shopper_S["数量"].values
          
        except IndexError:
          
          shopper_S_quantity = "0"
        
        print(shopper_S_quantity)
        #Mサイズ
        
        try:
          shopper_M = item_list[item_list["商品CD"].values == "9998998007"]
          shopper_M_quantity = shopper_M["数量"].values
          
        except IndexError:
          shopper_M_quantity = "0"
        
        print(shopper_M_quantity)
        #Lサイズ
        
        try :
          shopper_L = item_list[item_list["商品CD"].values == "9998998008"]
          shopper_L_quantity = shopper_L["数量"].values
          
        except IndexError:
          
          shopper_L_quantity = "0"

        
        print(shopper_L_quantity)  

        
        #wb = pyxl.load_workbook("C:/Users/fun-f/Desktop/ギフトショッパー.xlsx")#テスト用パス
        #wb = pyxl.load_workbook("C:/Users/fun-f/株式会社　ＴＲＩＮＩＴＹ　/遠藤 孝道 - 業務フォルダ/⓪共有業務/備品申請/備品2022年/2022 入力FORM【備品申請書】.xlsx")
        #wb = pyxl.load_workbook("C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/⓪共有業務/備品申請/備品2022年/2022 入力FORM【備品申請書】.xlsx")
        #wb = pyxl.load_workbook("C:/Users/fun-f\Downloads/2022 入力FORM【備品申請書】 (1).xlsx")
        wb = pyxl.load_workbook("C:/Users/fun-f/Downloads/2023 入力FORM【備品申請書】.xlsx")
        
   
        #ws = wb["適正在庫"]#テスト
        ws = wb["ギフトSP"]
        ws["A" + str(index_no)].value = shop_name[2]
        
        try :
          
          ws["E" + str(index_no)].value = shopper_S_quantity[0]
          
        except ValueError:  
          ws["E" + str(index_no)].value = 0
          
        except IndexError:
          
          ws["E" + str(index_no)].value = 0
            
          
          
        try:  
          ws["F" + str(index_no)].value = shopper_M_quantity[0]
          
        except ValueError:
          ws["F" + str(index_no)].value = 0 
          
        except IndexError:
          ws["F" + str(index_no)].value = 0 
          
          
        try:  
          ws["G" + str(index_no)].value = shopper_L_quantity[0]
        
        except  ValueError:
          
          ws["G" + str(index_no)].value =  0
          
        except  IndexError:
          
          ws["G" + str(index_no)].value =  0  
          
        index_no += 1
        
    #wb.save("C:/Users/fun-f/株式会社　ＴＲＩＮＩＴＹ　/遠藤 孝道 - 業務フォルダ/⓪共有業務/備品申請/備品2022年/2022 入力FORM【備品申請書】.xlsx")
        #wb.save("C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/⓪共有業務/備品申請/備品2022年/2022 入力FORM【備品申請書】.xlsx")
        wb.save("C:/Users/fun-f/Downloads/2023 入力FORM【備品申請書】.xlsx")

        
download()        

totalling()        
