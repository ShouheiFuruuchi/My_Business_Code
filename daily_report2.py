#このプログラムは店別品番別実績を自動ダウンロードを行う

#----------------------------------------------------------------------------------------------


import time
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ChromeOptions
import datetime
import os
import glob
import shutil
from operator import itemgetter
import tes
import datetime
import pandas as pd
import re
import numpy as np
import collections
import requests
import openpyxl as xlpy



#このプログラムは店別品番別実績を自動ダウンロードを行う


#ーーーーーーーーーー前回データの削除ーーーーーーーーーーーーー
folders = [0,1,2,3,4,5,6]
no = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]

w_day = datetime.datetime.today()

wd_no = w_day.weekday()#曜日Noを指定

main_dr = 'C:/Users/fun-f/Desktop/myfile'

print(wd_no)

to_file_path = str(main_dr) + '/' + str(wd_no)#drpathの指定

  #ーーーーーー曜日別商品実績ファイルクリアーーーーーーーーーー
  
if wd_no == 0:# 月曜日⇒0 火曜日⇒ 1 水曜日⇒ 2 木曜日⇒ 3 金曜日⇒ 4 土曜日⇒ 5 日曜日⇒ 6
  cl_sheet = pd.read_excel('C:/Users/fun-f/Desktop/myfile/クリアBOOK.xlsx')

  cl_df =pd.DataFrame(cl_sheet)
  for fd in folders:
    print(fd)
    for i in no:
      
      del_path = 'C:/Users/fun-f/Desktop/myfile/'+str(fd)+'/'+str(i)+'商品実績.xlsx'
      print(del_path)
      cl_df.to_excel(del_path)
      
  #ーーーーーーーーー実績ファイルクリアーーーーーーーーーーーーー
  
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/0/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/1/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/2/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/3/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/4/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/5/実績/実績.xlsx')
  cl_df.to_excel('C:/Users/fun-f/Desktop/myfile/6/実績/実績.xlsx')

  print('success')#削除完了
  
else:  
  print('Non Success!!')#削除ファイルなし

#ーーーーーーーー前回ダウンロードファイル削除ーーーーーーーーーー
dr_files = 'C:/Users/fun-f/Desktop/myfile/dataf'
dr_read = os.listdir(dr_files)

print(dr_read)

for file_name in dr_read:
  del_f_path = dr_files + '/' + file_name#削除ファイルパスの設定
  os.remove(del_f_path)#dataf内のファイルの削除
  
  
dr_files_2 = 'C:/Users/fun-f/Desktop/myfile/売上実績'  

dr_read_2 = os.listdir(dr_files_2)

print(dr_read_2)

for file_name_2 in dr_read_2:
  if file_name_2.endswith('.csv'):
    del_f_path2 = dr_files_2 + '/' + file_name_2#削除ファイルパスの設定
    os.remove(del_f_path2)#dataf内のファイルの削除
    
  
#ーーーーーーー今日の日付設定ーーーーーーーーー

fold = 'C:/Users/fun-f/Downloads'


todaytime = datetime.date.today()
tod = '{0:20%y%m%d}'.format(todaytime)#今日の日付(西暦)


#ーーーーーーー販売NETスクレイピングーーーーーーーーーーー

kasiwa = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[3]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[3]','柏','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[3]']
chiba = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[4]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[4]', '千葉','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[4]']
isesaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[9]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[9]','伊勢崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[9]']
nagamachi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[11]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[11]','長町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[11]']
hunabashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[12]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[12]','船橋','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[12]']
hujimi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[13]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[13]','富士見','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[13]']
reiku = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[15]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[15]','レイク','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[15]']
ebina = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[17]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[17]','海老名','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[17]']
musashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[18]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[18]','むさし','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[18]']
hiratuka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[19]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[19]','平塚','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[19]']
natori = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[20]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[20]','名取','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[20]']
otaka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[21]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[21]','大高','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[21]']
togocyo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[22]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[22]','東郷町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[22]']
ota = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[23]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[23]','太田','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[23]']
mito = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[24]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[24]','水戸','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[24]']
expo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[25]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[25]','EXPO','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[25]']
kawasaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[26]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[26]','川崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[26]']
sinmisato = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[27]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[27]','新三郷','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[27]']
makuhari = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[28]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[28]','幕張','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[28]']
kagamihara = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[29]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[29]','各務原','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[29]']

tenpo_list = [
  kasiwa,
  chiba,
  isesaki,
  nagamachi,
  hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara
  ]
  
tenpo_list_2 = [
  kasiwa,
  chiba,
  isesaki,
  nagamachi,
  hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara
  ]
  
weekday_writer_list = {
  0:"G",#月曜日
  1:"H",#火曜日
  2:"I",#水曜日
  3:"J",#木曜日
  4:"K",#金曜日
  5:"L",#土曜日
  6:"M" #日曜日
}



out_file2 = "C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/シフト管理/予実管理.xlsx"#予実管理出力ファイル

wb_out_file2 = xlpy.load_workbook(out_file2)

ws_out_file2 = wb_out_file2["実績データ"]

#---------------------------------------------------

url = 'http://tri.hanbai-net.com/system/Login.aspx'
#driver = webdriver.Chrome('C:/Users/fun-f/Downloads/chromedriver.exe')#旧
#driver = webdriver.Chrome('C:/Users/fun-f/Desktop/myfile/chromedriver.exe')
driver = webdriver.Chrome("C:/Users/fun-f/chromedriver.exe")#2021 0724

driver.get(url)

#id_1 = 'tenpo'
#id_2 = 'tenpo'

id_1 = 'trinityadmin'
id_2 = 'AdminTrinity'

loginid_1 = driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtUserCode"]')
loginid_2 = driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtPassword"]')

loginid_1.send_keys(id_1)#ユーザーIDを入力
loginid_2.send_keys(id_2)#パスワードを入力



driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnLogin"]').click() 
#ログインボタンをクリック

driver.get('http://tri.hanbai-net.com/system/00000000.aspx')

driver.find_element_by_xpath('//*[@id="Menu1"]/ul/li[7]').click()

#driver.find_element_by_xpath('//*[@id="Menu1:submenu:57"]/li[9]/a').click()
#'//*[@id="Menu1:submenu:58"]/li[9]/a'#変更前

driver.get('http://tri.hanbai-net.com/system/30021901.aspx?id=010199')


driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02"]').clear()#日付クリア

driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02"]').send_keys(str(tod))#日付入力

#----------全店------------

driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCondRun"]').click()#検索

time.sleep(5)

driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()#CSV出力

time.sleep(3)#一時待機

filelists = []
for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
    base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
    if ext == '.csv':#拡張子csvが一致した場合…
        if base == '品番売上集計':
            filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
            #print("file:{},csv:{}" .format(file,csv))
            filelists.sort(key=itemgetter(0), reverse=True)#
            MAX_CNT = 0
            for i, file in enumerate(filelists):
                if i > MAX_CNT-1:
                    print(file[0])
                    #file_1 = os.rename(i[0], 'kasi.csv')
                    os.rename(file[0], '全店.csv')
                    shutil.move('全店.csv','C:/Users/fun-f/Desktop/myfile/dataf')                        
time.sleep(1)                    

#--------店別---------
#★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
#品番別売上集計をダウンロード

for i_1 in tenpo_list:
  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownListCond04"]').click()#店舗名指定上段

  driver.find_element_by_xpath(str(i_1[0])).click()#店舗選択


  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownListCond05"]').click()#店舗名指定下段

  driver.find_element_by_xpath(str(i_1[1])).click()#店舗選択


  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCondRun"]').click()#検索

  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()#CSV出力

  time.sleep(3)#一時待機

  filelists = []
  for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
      base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
      if ext == '.csv':#拡張子csvが一致した場合…
          if base == '品番売上集計':
              filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
              #print("file:{},csv:{}" .format(file,csv))
              filelists.sort(key=itemgetter(0), reverse=True)#
              MAX_CNT = 0
              for i, file in enumerate(filelists):
                  if i > MAX_CNT-1:
                      print(file[0])
                      #file_1 = os.rename(i[0], 'kasi.csv')
                      os.rename(file[0], str(i_1[2]) + '.csv')
                      shutil.move(str(i_1[2]) + '.csv','C:/Users/fun-f/Desktop/myfile/dataf') 
                     
time.sleep(1)                    
        
print("SUCCESS!!")     
        
print("SUCCESS!!") 

#★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
#販売分析ログをダウンロード

driver.get('http://tri.hanbai-net.com/system/50010201.aspx?id=010199')#販売分析ログ



for i_1 in tenpo_list_2:
  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_DropDownList9"]').click()#店舗名指定上段

  driver.find_element_by_xpath(str(i_1[3])).click()#店舗選択


  driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()#CSV出力

  time.sleep(3)#一時待機

  filelists = []
  for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
      base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
      if ext == '.csv':#拡張子csvが一致した場合…
          if base == '販売分析ログ':
              filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
              #print("file:{},csv:{}" .format(file,csv))
              filelists.sort(key=itemgetter(0), reverse=True)#
              MAX_CNT = 0
              for i, file in enumerate(filelists):
                  if i > MAX_CNT-1:
                      print(file[0])
                      #file_1 = os.rename(i[0], 'kasi.csv')
                      os.rename(file[0], str(i_1[2]) + '販売ログ.csv')
                      shutil.move(str(i_1[2]) + '販売ログ.csv','C:/Users/fun-f/Desktop/myfile/dataf') 
                      
time.sleep(1)                      

print("SUCCESS")

#★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★

#売上集計*をダウンロード


driver.find_element_by_xpath('//*[@id="Menu1"]/ul/li[7]').click()

#driver.find_element_by_xpath('//*[@id="Menu1:submenu:57"]/li[14]/a').click()
#'//*[@id="Menu1:submenu:58"]/li[14]/a'#変更前


driver.get('http://tri.hanbai-net.com/system/30026401.aspx?id=010199')#売上集計＊

driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCondRun"]').click()

time.sleep(1)

driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnCSV"]').click()#CSV出力

time.sleep(3)#一時待機

filelists = []
for file in os.listdir("C:/Users/fun-f/Downloads"):#ディレクトリ内をfor文で取り出す
    base, ext = os.path.splitext(file)#splitextは拡張子を取得する関数
    if ext == '.csv':#拡張子csvが一致した場合…
        if base == '売上集計':
            filelists.append([file, os.path.getctime(file)])#filelistsに取り出したfileにダウンロード時間を追加
            #print("file:{},csv:{}" .format(file,csv))
            filelists.sort(key=itemgetter(0), reverse=True)#
            MAX_CNT = 0
            for i, file in enumerate(filelists):
                if i > MAX_CNT-1:
                    print(file[0])
                    #file_1 = os.rename(i[0], 'kasi.csv')
                    os.rename(file[0], '売上実績.csv')
                    shutil.move('売上実績.csv','C:/Users/fun-f/Desktop/myfile/売上実績')
                    
print("SUCCESS!!")      
                               
driver.close()




kasiwa = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[3]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[3]','柏','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[3]','FUN柏']

chiba = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[4]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[4]', '千葉','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[4]','FUN千葉C-one']

isesaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[9]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[9]','伊勢崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[9]','FUNスマーク伊勢崎']

nagamachi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[11]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[11]','長町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[11]','FUNララガーデン長町']

hunabashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[12]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[12]','船橋','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[12]','FUNららぽーとTOKYO-BAY']

hujimi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[13]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[13]','富士見','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[13]','FUNららぽーと富士見']

reiku = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[15]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[15]','レイク','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[15]','FUNイオンレイクタウン']

ebina = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[17]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[17]','海老名','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[17]','FUNららぽーと海老名']

musashi = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[18]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[18]','むさし','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[18]','FUNイオンモールむさし村山']

hiratuka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[19]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[19]','平塚','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[19]','FUNららぽーと湘南平塚']

natori = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[20]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[20]','名取','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[20]','FUNイオンモール名取']

otaka = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[21]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[21]','大高','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[21]','FUNイオンモール大高']

togocyo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[22]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[22]','東郷町','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[22]','FUNららぽーと愛知東郷']

ota = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[23]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[23]','太田','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[23]','FUNイオンモール太田']

mito = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[24]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[24]','水戸','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[24]','FUNイオンモール水戸内原']

expo = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[25]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[25]','EXPO','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[25]','FUNららぽーとEXPOCITY']

kawasaki = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[26]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[26]','川崎','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[26]','FUNラゾーナ川崎プラザ']

sinmisato = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[27]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[27]','新三郷','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[27]','FUNららぽーと新三郷']

makuhari = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[28]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[28]','幕張','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[28]','FUNイオンモール幕張新都心']

kagamihara = ['//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[29]','//*[@id="ContentPlaceHolder1_DropDownListCond05"]/option[29]','各務原','//*[@id="ContentPlaceHolder1_DropDownList9"]/option[29]','FUNイオンモール各務原']

tenpo_list = [
  kasiwa,
  chiba,
  isesaki,
  nagamachi,
  hunabashi,
  hujimi,
  reiku,
  ebina,
  musashi,
  hiratuka,
  natori,
  otaka,
  togocyo,
  ota,
  mito,
  expo,
  kawasaki,
  sinmisato,
  makuhari,
  kagamihara
  ]
  

tenpo_dic = {
  '柏':kasiwa,
  '千葉':chiba,
  '伊勢崎':isesaki,
  '長町':nagamachi,
  '船橋':hunabashi,
  '富士見':hujimi,
  'レイク':reiku,
  '海老名':ebina,
  'むさし':musashi,
  '平塚':hiratuka,
  '名取':natori,
  '大高':otaka,
  '東郷町':togocyo,
  '太田':ota,
  '水戸':mito,
  'EXPO':expo,
  '川崎':kawasaki,
  '新三郷':sinmisato,
  '幕張':makuhari,
  '各務原':kagamihara
}

daily_list = {  
  0:["M","O","P"],
  1:["Q","S","T"],
  2:["U","W","X"],
  3:["Y","AA","AB"],
  4:["AC","AE","AF"],
  5:["AG","AI","AJ"],
  6:["AK","AM","AN"]
}


#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#予実管理ファイル

#一致店舗の列調整
tenpo_pitch = {
  '柏':0,
  '千葉':1,
  '伊勢崎':2,
  '長町':3,
  '船橋':4,
  '富士見':5,
  'レイク':6,
  '海老名':7,
  'むさし':8,
  '平塚':9,
  '名取':10,
  '大高':11,
  '東郷町':12,
  '太田':13,
  '水戸':14,
  'EXPO':15,
  '川崎':16,
  '新三郷':17,
  '幕張':18,
  '各務原':19
}


#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

    
file_path = 'C:/Users/fun-f/Desktop/myfile/dataf'
  
file_names = os.listdir(file_path)


files_list = []
concat_file = []


for i in file_names:
  if "販売ログ" in i:
    files_list.append(i)

#顧客データ分析

range_count = len(files_list) #- 1

folder_path = 'C:/Users/fun-f/Desktop/analysis/create_file/'

file_name_1 = os.listdir(folder_path)


#bk_1 = xlpy.load_workbook("C:/Users/fun-f/Desktop/analysis/週間分析途中.xlsx")#当日実績を入力するファイル
bk_1 = xlpy.load_workbook(folder_path + file_name_1[0])#当日実績を入力するファイル
#------------------------------------------------------------------------------
for file_n,n in zip(files_list,range(0,int(range_count))) :

  
  select_file = file_n
  
  r_file = pd.read_csv(file_path + '/' + select_file,encoding="SHIFT-JIS")#販売分析ログ
  r_file2 = pd.read_csv("C:/Users/fun-f/Desktop/myfile/売上実績/売上実績.csv",encoding="SHIFT-JIS")#売上集計
  
  df_r_file = pd.DataFrame(r_file)
  df_r_file2 = pd.DataFrame(r_file2)
  
  order_n = pd.DataFrame(df_r_file["伝票番号"],columns=["伝票番号"])
  item_cd = pd.DataFrame(df_r_file["商品コード"].astype('str').str.zfill(10).str[:10].values,columns=["商品CD"])
  item_name = pd.DataFrame(df_r_file["商品名"],columns=["商品名"])
  category_cd = pd.DataFrame(df_r_file["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
  quantity = pd.DataFrame(df_r_file["伝票明細数量"].values,columns=["数量"])
  amount = pd.DataFrame(df_r_file["伝票明細小計金額"].values,columns=["金額"])
  
  set_data = pd.concat([order_n,item_cd,item_name,category_cd,quantity,amount],axis=1)
    
  #filter_data = set_data[set_data["金額"] >= 100]
  
  filter_1 = set_data[set_data["アイテムCD"] != "98"] #ショッパー除外
  
  filter_2 = filter_1[filter_1["アイテムCD"] != "14"] #サンプル除外
  
  filter_data = filter_2[filter_2["金額"] != 50] #マスク除外
  
  
  shop = file_n.replace('販売ログ.csv','')#店名
  
  ws_1 = bk_1[str(shop)]
  
 

  #for shop_n in df_r_file2["拠点名"]:
    
  select_shop = df_r_file2[df_r_file2["拠点名"] == tenpo_dic[str(shop)][4]]
  
  buget = select_shop["売上予算"].values[0]
      
  
  number_of_customer = len(np.unique(filter_data["伝票番号"]))#客数(※ショッパーのみなし)
  filter_quantity = sum(filter_data["数量"])#数量(※ショッパーなし)
  order_count = collections.Counter(filter_data["伝票番号"].values).values()
  filter_sales = sum(filter_data["金額"])
  try :
    set_ratio = '{: .2f}'.format(float(filter_quantity / number_of_customer)) #SET率(※ショッパーなし)
    sales_ratio = '{: .1f}'.format(filter_sales / buget * 100) 
    
  except ZeroDivisionError:
    
    set_ratio = '0' #SET率(※ショッパーなし)
    sales_ratio = '0'
  
  count = 0
  for i in order_count:
    if i > 1:
      count += 1
      
  set_quantity = count #SET数
  
  try:
    set_quantity_ratio = '{: .1f}'.format(float(set_quantity / number_of_customer * 100))
    
  except ZeroDivisionError:
    set_quantity_ratio = '0'


  #print(shop)
  #print("売上予算 ⇒ " + str(buget))
  #print("売上実績 ⇒ " + str(filter_sales))
  #print("客数 ⇒ " + str(number_of_customer))
  #print("数量 ⇒ " + str(filter_quantity))
  #print(set_ratio)
  #print(set_quantity)
  #print(set_quantity_ratio)
  
  
  
  data1 = pd.concat([order_n,item_cd,item_name,category_cd,quantity,amount],axis=1)

  op = data1[data1['アイテムCD'] == '01']
  cd = data1[data1['アイテムCD'] == '02']
  jk = data1[data1['アイテムCD'] == '03']
  kt = data1[data1['アイテムCD'] == '04']
  cs = data1[data1['アイテムCD'] == '05']
  ct = data1[data1['アイテムCD'] == '06']
  bl = data1[data1['アイテムCD'] == '07']
  sk = data1[data1['アイテムCD'] == '08']
  pt = data1[data1['アイテムCD'] == '09']
  tr = data1[data1['アイテムCD'] == '10']
  inn = data1[data1['アイテムCD'] == '11']
  setup = data1[data1['アイテムCD'] == '12']
  acc = data1[data1['アイテムCD'] == '13']
  sh = data1[data1['アイテムCD'] == '14']

  #print(shop + "⇒" +str("{: .1f}".format((len(np.unique(op['伝票番号'].values))/number_of_customer) * 100)) + "%")


  #アイテム金額構成比リスト

  
  try :
    df_op = pd.DataFrame({'アイテムCD':['OP'],'CD':['01'],'用途分類':['OP/SET UP'],'用途CD':['1'],'金額':[sum(op['金額'].values)],'売上構成比':[sum(op['金額'].values/filter_sales)],'購入比':[len(np.unique(op['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    df_op = pd.DataFrame({'アイテムCD':['OP'],'CD':['01'],'用途分類':['OP/SET UP'],'用途CD':['1'],'金額':[0],'売上構成比':[0],'購入比':[0]})
   
  try :
    df_cd = pd.DataFrame({'アイテムCD':['CD'],'CD':['02'],'用途分類':['羽織'],'用途CD':['2'],'金額':[sum(cd['金額'].values)],'売上構成比':[sum(cd['金額'].values/filter_sales)],'購入比':[len(np.unique(cd['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_cd = pd.DataFrame({'アイテムCD':['CD'],'CD':['02'],'用途分類':['羽織'],'用途CD':['2'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try:
    df_jk = pd.DataFrame({'アイテムCD':['JK'],'CD':['03'],'用途分類':['羽織'],'用途CD':['2'],'金額':[sum(jk['金額'].values)],'売上構成比':[sum(jk['金額'].values/filter_sales)],'購入比':[len(np.unique(jk['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError :
    
    df_jk = pd.DataFrame({'アイテムCD':['JK'],'CD':['03'],'用途分類':['羽織'],'用途CD':['2'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try :  
    df_kt = pd.DataFrame({'アイテムCD':['KT'],'CD':['04'],'用途分類':['トップス'],'用途CD':['3'],'金額':[sum(kt['金額'].values)],'売上構成比':[sum(kt['金額'].values/filter_sales)],'購入比':[len(np.unique(kt['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    df_kt = pd.DataFrame({'アイテムCD':['KT'],'CD':['04'],'用途分類':['トップス'],'用途CD':['3'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try :
    df_cs = pd.DataFrame({'アイテムCD':['CS'],'CD':['05'],'用途分類':['トップス'],'用途CD':['3'],'金額':[sum(cs['金額'].values)],'売上構成比':[sum(cs['金額'].values/filter_sales)],'購入比':[len(np.unique(cs['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError :
    
    df_cs = pd.DataFrame({'アイテムCD':['CS'],'CD':['05'],'用途分類':['トップス'],'用途CD':['3'],'金額':[0],'売上構成比':[0],'購入比':[0]})
  
  try :  
    df_ct = pd.DataFrame({'アイテムCD':['CT'],'CD':['06'],'用途分類':['羽織'],'用途CD':['2'],'金額':[sum(ct['金額'].values)],'売上構成比':[sum(ct['金額'].values/filter_sales)],'購入比':[len(np.unique(ct['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_ct = pd.DataFrame({'アイテムCD':['CT'],'CD':['06'],'用途分類':['羽織'],'用途CD':['2'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try :  
    
    df_bl = pd.DataFrame({'アイテムCD':['BL'],'CD':['07'],'用途分類':['トップス'],'用途CD':['3'],'金額':[sum(bl['金額'].values)],'売上構成比':[sum(bl['金額'].values/filter_sales)],'購入比':[len(np.unique(bl['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_bl = pd.DataFrame({'アイテムCD':['BL'],'CD':['07'],'用途分類':['トップス'],'用途CD':['3'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try :  
    
    df_sk = pd.DataFrame({'アイテムCD':['SK'],'CD':['08'],'用途分類':['ボトムス'],'用途CD':['4'],'金額':[sum(sk['金額'].values)],'売上構成比':[sum(sk['金額'].values/filter_sales)],'購入比':[len(np.unique(sk['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_sk = pd.DataFrame({'アイテムCD':['SK'],'CD':['08'],'用途分類':['ボトムス'],'用途CD':['4'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
    
  try:    
    df_pt = pd.DataFrame({'アイテムCD':['PT'],'CD':['09'],'用途分類':['ボトムス'],'用途CD':['4'],'金額':[sum(pt['金額'].values)],'売上構成比':[sum(pt['金額'].values/filter_sales)],'購入比':[len(np.unique(pt['伝票番号'].values))/number_of_customer]})
    
    
  except ZeroDivisionError:
    
    df_pt = pd.DataFrame({'アイテムCD':['PT'],'CD':['09'],'用途分類':['ボトムス'],'用途CD':['4'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
  try :  
    df_tr = pd.DataFrame({'アイテムCD':['TR'],'CD':['10'],'用途分類':['トップス'],'用途CD':['3'],'金額':[sum(tr['金額'].values)],'売上構成比':[sum(tr['金額'].values/filter_sales)],'購入比':[len(np.unique(tr['伝票番号'].values))/number_of_customer]})
  except ZeroDivisionError: 
    
    df_tr = pd.DataFrame({'アイテムCD':['TR'],'CD':['10'],'用途分類':['トップス'],'用途CD':['3'],'金額':[0],'売上構成比':[0],'購入比':[0]}) 
    
    
  try :  
    df_inn = pd.DataFrame({'アイテムCD':['INN'],'CD':['11'],'用途分類':['インナー'],'用途CD':['5'],'金額':[sum(inn['金額'].values)],'売上構成比':[sum(inn['金額'].values/filter_sales)],'購入比':[len(np.unique(inn['伝票番号'].values))/number_of_customer]})
  except ZeroDivisionError:  
    
    df_inn = pd.DataFrame({'アイテムCD':['INN'],'CD':['11'],'用途分類':['インナー'],'用途CD':['5'],'金額':[0],'売上構成比':[0],'購入比':[0]})
  
  try:  
    df_setup = pd.DataFrame({'アイテムCD':['SETUP'],'CD':['12'],'用途分類':['OP/SET UP'],'用途CD':['1'],'金額':[sum(setup['金額'].values)],'売上構成比':[sum(setup['金額'].values/filter_sales)],'購入比':[len(np.unique(setup['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_setup = pd.DataFrame({'アイテムCD':['SETUP'],'CD':['12'],'用途分類':['OP/SET UP'],'用途CD':['1'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
    
  try:  
    df_acc = pd.DataFrame({'アイテムCD':['ACC'],'CD':['13'],'用途分類':['ACC'],'用途CD':['6'],'金額':[sum(acc['金額'].values)],'売上構成比':[sum(acc['金額'].values/filter_sales)],'購入比':[len(np.unique(acc['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_acc = pd.DataFrame({'アイテムCD':['ACC'],'CD':['13'],'用途分類':['ACC'],'用途CD':['6'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
    
  try :  
    df_sh = pd.DataFrame({'アイテムCD':['SH'],'CD':['14'],'用途分類':['シューズ'],'用途CD':['7'],'金額':[sum(sh['金額'].values)],'売上構成比':[sum(sh['金額'].values/filter_sales)],'購入比':[len(np.unique(sh['伝票番号'].values))/number_of_customer]})
    
  except ZeroDivisionError:
    
    df_sh = pd.DataFrame({'アイテムCD':['SH'],'CD':['14'],'用途分類':['シューズ'],'用途CD':['7'],'金額':[0],'売上構成比':[0],'購入比':[0]})
    
    
    

  item_rank = pd.concat([df_op,df_cd,df_jk,df_kt,df_cs,df_ct,df_bl,df_sk,df_pt,df_tr,df_inn,df_setup,df_acc,df_sh],axis=0)
  
  #★店別アイテム実績を出力
  for cont_no in range(0,13):#12 ⇒ 13 変更
  
    ws_1[str(daily_list[int(wd_no)][0]) + str(17 + int(cont_no))].value = item_rank.values[cont_no][4]
    ws_1[str(daily_list[int(wd_no)][1]) + str(17 + int(cont_no))].value = item_rank.values[cont_no][5]
    ws_1[str(daily_list[int(wd_no)][2]) + str(17 + int(cont_no))].value = item_rank.values[cont_no][6]
  
  
  ws_1[str(daily_list[int(wd_no)][0]) + str(11)].value = buget#売上予算
  ws_1[str(daily_list[int(wd_no)][0]) + str(12)].value = filter_sales#売上実績
  ws_1[str(daily_list[int(wd_no)][0]) + str(13)].value = number_of_customer#客数
  ws_1[str(daily_list[int(wd_no)][0]) + str(14)].value = filter_sales / number_of_customer#客単価
  ws_1[str(daily_list[int(wd_no)][0]) + str(15)].value = set_ratio#P率
       
  
  #合計欄  
  ws_1[str(daily_list[int(wd_no)][0]) + str(30)].value = sum(item_rank["金額"].values) - item_rank.values[13][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(30)].value = sum(item_rank["売上構成比"].values) - item_rank.values[13][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(30)].value = sum(item_rank["購入比"].values) - item_rank.values[13][6]
  
  #★店別用途区分出力
  #OP/SETUP
  ws_1[str(daily_list[int(wd_no)][0]) + str(33)].value = item_rank.values[0][4] + item_rank.values[11][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(33)].value = item_rank.values[0][5] + item_rank.values[11][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(33)].value = item_rank.values[0][6] + item_rank.values[0][6]
  
  #TOPS
  ws_1[str(daily_list[int(wd_no)][0]) + str(34)].value = item_rank.values[3][4] + item_rank.values[4][4] + item_rank.values[6][4] + item_rank.values[9][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(34)].value = item_rank.values[3][5] + item_rank.values[4][5] + item_rank.values[6][5] + item_rank.values[9][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(34)].value = item_rank.values[3][6] + item_rank.values[4][6] + item_rank.values[6][6] + item_rank.values[9][6]
  
  #BOTTOMS
  ws_1[str(daily_list[int(wd_no)][0]) + str(35)].value = item_rank.values[7][4] + item_rank.values[8][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(35)].value = item_rank.values[7][5] + item_rank.values[8][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(35)].value = item_rank.values[7][6] + item_rank.values[8][6]

  #羽織
  ws_1[str(daily_list[int(wd_no)][0]) + str(36)].value = item_rank.values[1][4] + item_rank.values[2][4] +item_rank.values[5][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(36)].value = item_rank.values[1][5] + item_rank.values[2][5] +item_rank.values[5][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(36)].value = item_rank.values[1][6] + item_rank.values[2][6] +item_rank.values[5][6]
  
  #インナー
  ws_1[str(daily_list[int(wd_no)][0]) + str(37)].value = item_rank.values[10][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(37)].value = item_rank.values[10][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(37)].value = item_rank.values[10][6]
  
  #ACC
  ws_1[str(daily_list[int(wd_no)][0]) + str(38)].value = item_rank.values[12][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(38)].value = item_rank.values[12][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(38)].value = item_rank.values[12][6]
  
  #合計欄
  ws_1[str(daily_list[int(wd_no)][0]) + str(39)].value = sum(item_rank["金額"].values) - item_rank.values[13][4]
  ws_1[str(daily_list[int(wd_no)][1]) + str(39)].value = sum(item_rank["売上構成比"].values) - item_rank.values[13][5]
  ws_1[str(daily_list[int(wd_no)][2]) + str(39)].value = sum(item_rank["購入比"].values) - item_rank.values[13][6]
  
  bk_1.save(folder_path + file_name_1[0])
  
 
  

  
  set_data_comp = pd.DataFrame([{"店舗名":shop,"売上予算":buget,"売上実績":filter_sales,"予算達成率": sales_ratio,"客数":number_of_customer,"点数":filter_quantity,"P率":set_ratio,"SET販売数":set_quantity,"SET比率":set_quantity_ratio}])
  
  #set_data_comp = pd.DataFrame([{"店舗名":shop,buget,filter_sales,number_of_customer,filter_quantity,set_ratio,set_quantity,set_quantity_ratio}])
  
  concat_file.append(set_data_comp)
  
  #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
  
  w_day_2 = '{:%Y%m%d}'.format(datetime.datetime.today())
  year_2 = w_day_2[0:4]
  month_2 = w_day_2[4:6]
  #day_2 = int(w_day_2[6:8]) 
  day_2 = w_day_2[6:8]
  
  #day_2 = str(day_2).zfill(2)



  #select_day = "2022-4-14"

  select_day = str(year_2) + "-" + str(month_2) +"-" + str(day_2) + " 00:00:00"
  #select_day = str(year_2) + "/" + str(month_2) +"/" + str(day_2)
  
  print(select_day)
  print("ここ" + str(select_day))
  for i in range(0,366):#364
    target_cell = ws_out_file2["A" + str(4 + i)].value
    
    if str(target_cell) == str(select_day):
      print("YES")
      
      cell_r = 4 + i
      print(cell_r)
  
  
      pitch = 7 #列間隔
      out_file2_header = cell_r#行番号
      out_file2_col =2 + (tenpo_pitch[shop] * pitch)#列番号


      cell_1 = ws_out_file2.cell(out_file2_header,out_file2_col).value = buget#予算
      cell_2 = ws_out_file2.cell(out_file2_header,out_file2_col + 1).value = filter_sales#実績
      cell_3 = ws_out_file2.cell(out_file2_header,out_file2_col + 2).value = number_of_customer#客数
      cell_4 = ws_out_file2.cell(out_file2_header,out_file2_col + 3).value = filter_quantity#点数
      cell_5 = ws_out_file2.cell(out_file2_header,out_file2_col + 4).value = set_ratio#P率
      cell_6 = ws_out_file2.cell(out_file2_header,out_file2_col + 5).value = filter_sales / number_of_customer#客単価
      cell_7 = ws_out_file2.cell(out_file2_header,out_file2_col + 6).value = set_quantity_ratio#SET比

      print(cell_1)
      
    else:
      
      print(target_cell)    

  wb_out_file2.save(out_file2)
  
  #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

  
  
comp_file = pd.concat(concat_file)    

print(comp_file.sort_values('P率',ascending=False))

comp_file = comp_file.sort_values('P率',ascending=False)
#-------------------------------------------------------------------

print(comp_file)

p_1 = comp_file.values[0]
p_2 = comp_file.values[1]
p_3 = comp_file.values[2]
p_4 = comp_file.values[3]
p_5 = comp_file.values[4]
p_6 = comp_file.values[5]
p_7 = comp_file.values[6]
p_8 = comp_file.values[7]
p_9 = comp_file.values[8]
p_10 = comp_file.values[9]
p_11 = comp_file.values[10]
p_12 = comp_file.values[11]
p_13 = comp_file.values[12]
p_14 = comp_file.values[13]
p_15 = comp_file.values[14]
p_16 = comp_file.values[15]
p_17 = comp_file.values[16]
p_18 = comp_file.values[17]
p_19 = comp_file.values[18]
p_20 = comp_file.values[19]


data_list = [
  p_1,p_2,p_3,p_4,p_5,p_6,p_7,p_8,p_9,p_10,p_11,p_12,p_13,p_14,p_15,p_16,p_17,p_18,p_19,p_20
             ]


#--------------------------------------------------------
w_day = datetime.datetime.today()

#w_day_str = '{: /}'.format(w_day)
w_day_str = "{}/{}/{}".format(w_day.year, w_day.month, w_day.day)
#アウトプットファイルの指定
stock_path = "C:/Users/fun-f/Desktop/実績集計.xlsx"
#stock_path = "C:/Users/fun-f/Downloads/【20211221】 2022 3月シフト 【販売部】 ver 14.xlsm"


wb = xlpy.load_workbook(stock_path)
ws = wb["Sheet1"]
#ws = wb["実績"]

last_low = ws.max_row + 1
print(last_low)
range_count = int(len(data_list) + 1)
#--------------------------------------------------------
for sheet_no,data_no in zip(range(0,range_count),data_list):
  ws["A" + str(last_low+sheet_no)].value = w_day_str
  ws["B" + str(last_low+sheet_no)].value = data_no[0]#店舗名
  ws["C" + str(last_low+sheet_no)].value = int(data_no[1])#売上予算
  ws["D"+ str(last_low+sheet_no)].value = int(data_no[2])#売上実績
  ws["E"+ str(last_low+sheet_no)].value = float(data_no[3])#予算達成率
  ws["F"+ str(last_low+sheet_no)].value = int(data_no[4])#客数
  ws["G"+ str(last_low+sheet_no)].value = float(data_no[6])#P率
  ws["H"+ str(last_low+sheet_no)].value = int(data_no[7])#SET販売数
  ws["I"+ str(last_low+sheet_no)].value = float(data_no[8])#SET比
  
  #★★★★★★★★
  #wb.save(stock_path)
  #wb.close()
  #★★★★★★★★
  
#TOKEN = 'NxPDQg0tpI4oY6BZHo7vkZ3gxPtTIijpWFyN85xL2q1'#テストの部屋トークン
TOKEN = 'TNKXBcpEMmK4JAmRPaOyVABkA5GWIJkIHQOnsyfu4MD'#FUNの部屋トークン
api_url = 'https://notify-api.line.me/api/notify'
headers = {'Authorization' : 'Bearer ' + TOKEN}
#message = ('\n'+'柏'+'\n'+'【売上予算/実績】'+'\n' + str(mg1) +'\n' +'【P率】' +str(p1) +'\n'+ '【客数】'+ str(noc_2) +str(p2)+str(p3))
message_1 = ('\n'+'今日のP率(※ショッパー抜き)'+'\n'+
             "1位〜10位"+'\n'+'\n'+
           
           '1位 ' + str(p_1[0]) + '\n' +
           ' P率' + str(p_1[6]) + '/ SET比'+ 
           (p_1[8]) + '% ' + '\n' +
           '実績 ¥' + str(p_1[2]) + ' 達成率' + str(p_1[3]) + '% ' + '\n' +"\n" +
            
           "2位 " + str(p_2[0]) + "\n" +
           ' P率' + str(p_2[6]) + '/ SET比'+ str(p_2[8])+ "% " + "\n" +
           "実績 ¥" + str(p_2[2]) + ' 達成率' +str(p_2[3])+ "% " + "\n" +"\n" +
           
           "3位 " + str(p_3[0]) + "\n" +
           ' P率' + str(p_3[6]) + '/ SET比'+ str(p_3[8])+ "% " + "\n" +
           "実績 ¥" + str(p_3[2]) + ' 達成率' +str(p_3[3])+ "% " + "\n" +"\n" +
           
           "4位 " + str(p_4[0]) + "\n" +
           ' P率' + str(p_4[6]) + '/ SET比'+ str(p_4[8])+ "% " + "\n" +
           "実績 ¥" + str(p_4[2]) + ' 達成率' +str(p_4[3])+ "% " + "\n" +"\n" +
           
           "5位 " + str(p_5[0]) + "\n" +
           ' P率' + str(p_5[6]) + '/ SET比'+ str(p_5[8])+ "% " + "\n" +
           "実績 ¥" + str(p_5[2]) + ' 達成率' +str(p_5[3])+ "% " + "\n" +"\n" +
           
           "6位 " + str(p_6[0]) + "\n" +
           
           ' P率' + str(p_6[6]) + '/ SET比'+ str(p_6[8])+ "% " + "\n" +
           "実績 ¥" + str(p_6[2]) + ' 達成率' +str(p_6[3])+ "% " + "\n" +"\n" +
           
           "7位 " + str(p_7[0]) + "\n" +
           ' P率' + str(p_7[6]) + '/ SET比'+ str(p_7[8])+ "% " + "\n" +
           "実績 ¥" + str(p_7[2]) + ' 達成率' +str(p_7[3])+ "% " + "\n" +"\n" +
           
           "8位 " + str(p_8[0]) + "\n" +
           ' P率' + str(p_8[6]) + '/ SET比'+ str(p_8[8])+ "% " + "\n" +
           "実績 ¥" + str(p_8[2]) + ' 達成率' +str(p_8[3])+ "% " + "\n" +"\n" +
           
           "9位 " + str(p_9[0]) + "\n" +
           ' P率' + str(p_9[6]) + '/ SET比'+ str(p_9[8])+ "% " + "\n" +
           "実績 ¥" + str(p_9[2]) + ' 達成率' +str(p_9[3])+ "% " + "\n" +"\n" +
           
           "10位 " + str(p_10[0]) + "\n" +
           ' P率' + str(p_10[6]) + '/ SET比'+ str(p_10[8])+ "% " + "\n" +
           "実績 ¥" + str(p_10[2]) + ' 達成率' +str(p_10[3])+ "% " + "\n" 
           
)
message_2 = ('\n'+'今日のP率(※ショッパー抜き)'+'\n'+
             "11位〜20位"+'\n'+'\n'+
                        
           "11位 " + str(p_11[0]) + "\n" +
           ' P率' + str(p_11[6]) + '/ SET比'+ str(p_11[8])+ "% " + "\n" +
           "実績 ¥" + str(p_11[2]) + ' 達成率' +str(p_11[3])+ "% " + "\n" +"\n" +
           
           "12位 " + str(p_12[0]) + "\n" +
           ' P率' + str(p_12[6]) + '/ SET比'+ str(p_12[8])+ "% " + "\n" +
           "実績 ¥" + str(p_12[2]) + ' 達成率' +str(p_12[3])+ "% " + "\n" +"\n" +
           
           "13位 " + str(p_13[0]) + "\n" +
           ' P率' + str(p_13[6]) + '/ SET比'+ str(p_13[8])+ "% " + "\n" +
           "実績 ¥" + str(p_13[2]) + ' 達成率' +str(p_13[3])+ "% " + "\n" +"\n" +
           
           "14位 " + str(p_14[0]) + "\n" +
           ' P率' + str(p_14[6]) + '/ SET比'+ str(p_14[8])+ "% " + "\n" +
           "実績 ¥" + str(p_14[2]) + ' 達成率' +str(p_14[3])+ "% " + "\n" +"\n" +
           
           "15位 " + str(p_15[0]) + "\n" +
           ' P率' + str(p_15[6]) + '/ SET比'+ str(p_15[8])+ "% " + "\n" +
           "実績 ¥" + str(p_15[2]) + ' 達成率' +str(p_15[3])+ "% " + "\n" +"\n" +
           
           "16位 " + str(p_16[0]) + "\n" +
           ' P率' + str(p_16[6]) + '/ SET比'+ str(p_16[8])+ "% " + "\n" +
           "実績 ¥" + str(p_16[2]) + ' 達成率' +str(p_16[3])+ "% " + "\n" +"\n" +
           
            "17位 " + str(p_17[0]) + "\n" +
           ' P率' + str(p_17[6]) + '/ SET比'+ str(p_17[8])+ "% " + "\n" +
           "実績 ¥" + str(p_17[2]) + ' 達成率' + str(p_17[3])+ "% " + "\n" +"\n" +
           
           "18位 " + str(p_18[0]) + "\n" +
           ' P率' + str(p_18[6]) + '/ SET比'+ str(p_18[8])+ "% " + "\n" +
           "実績 ¥" + str(p_18[2]) + ' 達成率' +str(p_18[3])+ "% " + "\n" +"\n" +
           
           "19位 " + str(p_19[0]) + "\n" +
           ' P率' + str(p_19[6]) + '/ SET比'+ str(p_19[8])+ "% " + "\n" +
           "実績 ¥" + str(p_19[2]) + ' 達成率' +str(p_19[3])+ "% " + "\n" +"\n" +
           
           "20位 " + str(p_20[0]) + "\n" +
           ' P率' + str(p_20[6]) + '/ SET比'+ str(p_20[8])+ "% " + "\n" +
           "実績 ¥" + str(p_20[2]) + ' 達成率' +str(p_20[3])+ "% " + "\n" +
      
            '\n'+'質問や不明点あれば古内までご連絡下さい！'+'\n'+'\n'+'よろしくお願いいたします。')
#(+'\n'+'岐阜'+str(p5)+'\n'+'長町'+str(p6)+'\n'+'船橋'+str(p7)+'\n'+'富士見'+str(p8)+'\n'+'レイク'+str(p9)+'\n'+'海老名')
#(+str(p10)+'\n'+'むさし'+str(p11)+'\n'+'平塚'+str(p12)+'\n'+'名取'+str(p13)+'\n'+'大高'+str(p14)+'\n'+'東郷町'+str(p15)+'\n'+'太田'+str(p16)+'\n'+'水戸'+str(p17)+'\n'+'EXPO'+str(p18)+'\n'+'川崎'+str(p19)+'\n'+'新三郷'+str(p20)+'\n'+'詳細はOneDriveの【シフト管理】売上実績ファイルを参照下さい！')
payload = {'message': message_1}

requests.post(api_url, headers=headers, params=payload)   
#print("SUCCESSFULL!!")


payload = {'message': message_2}

requests.post(api_url, headers=headers, params=payload)   
print("SUCCESSFULL!!")



#xlpy.load_workbook()