import pandas as pd
import openpyxl as pyxl
import os
import numpy as np
import requests
import datetime 

tenpo = [
  
    ["01001008 FUN柏","柏"],
    ["01001009 FUN千葉C-one","千葉"],
    ["01001028 FUNスマーク伊勢崎","伊勢崎"],
    ["01001034 FUNららぽーと富士見","富士見"],
    ["01001036 FUNイオンレイクタウン","レイク"],
    ["01001038 FUNららぽーと海老名","海老名"],
    ["01001039 FUNイオンモールむさし村山","むさし"],
    ["01001040 FUNららぽーと湘南平塚","平塚"],
    ["01001041 FUNイオンモール名取","名取"],
    ["01001042 FUNイオンモール大高","大高"],
    ["01001043 FUNららぽーと愛知東郷","東郷町"],
    ["01001044 FUNイオンモール太田","太田"],
    ["01001045 FUNイオンモール水戸内原","水戸"],
    ["01001046 FUNららぽーとEXPOCITY","EXPO"],
    ["01001047 FUNラゾーナ川崎プラザ","川崎"],
    ["01001048 FUNららぽーと新三郷","新三郷"],
    ["01001049 FUNイオンモール幕張新都心","幕張"],
    ["01001050 FUNイオンモール各務原","各務原"],
    ["01001051 FUNららぽーと堺","堺"],
    
]

#用途別ＣＤ
division_CD = {
  "01":"OP/SETUP",                                                                                            
  "02":"羽織",
  "03":"羽織",
  "04":"TOPs",
  "05":"TOPs",
  "06":"羽織",
  "07":"TOPs",
  "08":"BOTTOMs",
  "09":"BOTTOMs",
  "10":"TOPs",
  "11":"INN",
  "12":"OP/SETUP",
  "13":"ACC",
  "15":"SH",
  
}



item_cd_list = {
  "OP":"01",                                                                                            
  "CD":"02",
  "JK":"03",
  "KT":"04",
  "CS":"05",
  "CT":"06",
  "BL":"07",
  "SK":"08",
  "PT":"09",
  "TR":"10",
  "INN":"11",
  "SETUP":"12",
  "ACC":"13",
  "SH":"15",
}            
       
       
item_underprice_list = {
  "01":2299,                                                                                            
  "02":1799,
  "03":2599,
  "04":1599,
  "05":1599,
  "06":2599,
  "07":1799,
  "08":1999,
  "09":1999,
  "10":1799,
  "11":1299,
  "12":2599,
  "13":799,
  "15":2599,
}               

#日付と取得
today = datetime.datetime.today()
d_count = int(today.day)
day_element = datetime.timedelta(days=d_count)
yestaday = today - day_element
y = today.year
m = today.month
d = today.day

ToDay1 = str(y) + "/" + str(m) + "/" + str(d)
ToDay2 = str(y) + "年" + str(m) + "月" + str(d) + "日"

print(str(y) + "/" + str(m) + "/" + str(d))



file_path_all = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/myfile/dataf2/全店販売伝票明細.csv"#販売分析ログ

r_file_all = pd.read_csv(file_path_all,encoding="cp932")
df_r_file_all = pd.DataFrame(r_file_all)


order_n = pd.DataFrame(df_r_file_all["伝票番号"],columns=["伝票番号"])
day = pd.DataFrame(df_r_file_all["営業日付"],columns=["営業日付"])
item_cd = pd.DataFrame(df_r_file_all["商品コード"].astype('str').str.zfill(10).str[:10].values,columns=["商品CD"])
item_name = pd.DataFrame(df_r_file_all["商品名"],columns=["商品名"])
category_cd = pd.DataFrame(df_r_file_all["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
quantity = pd.DataFrame(df_r_file_all["数量"].values,columns=["数量"])
amount = pd.DataFrame(df_r_file_all["小計金額"].values,columns=["金額"])
cost = pd.DataFrame(df_r_file_all["原価"].values,columns=["原価"])#原価
shop_name = pd.DataFrame(df_r_file_all["店舗名"].values,columns=["店舗名"])

set_data_all = pd.concat([order_n,day,item_cd,item_name,category_cd,quantity,amount,cost,shop_name],axis=1)
  
#filter_data = set_data[set_data["金額"] >= 100]

filter_1_all = set_data_all[set_data_all["アイテムCD"] != "98"] #ショッパー除外

filter_2_all = filter_1_all[filter_1_all["アイテムCD"] != "14"] #サンプル除外

#追加コード
filter_3_all = filter_2_all[filter_2_all["金額"] > 0 ]#返品データを削除

filter_data_all = filter_3_all[filter_3_all["金額"] != 50] #マスク除外

#差し引きデータ
Pull_data_all = filter_2_all[filter_2_all["金額"] < 0 ]


CustCD_list_all = np.unique(filter_data_all["伝票番号"].values)
Price_list_all = np.unique(filter_data_all["金額"].values)      

print("金額リスト", Price_list_all)

#客単価ランキング
def UnitPrice():
  #====================================================
  #単日客単価ランキングを作成
  #====================================================
  
  Unit_Rank_list = []
  OUT_FILE_PATH = "C:/Users/古内翔平/Desktop/客単価ランキング.xlsx"

  
  wb_UP = pyxl.load_workbook(OUT_FILE_PATH)
  ws_UP = wb_UP[str(7) + "月"]  
  ws_UP2 = wb_UP["DataBase"] 
  for C_CD in CustCD_list_all:
    select_data = filter_data_all[filter_data_all["伝票番号"] == C_CD ]
    select_data_shop = select_data["店舗名"].values[0]
    select_data_UP = sum(select_data["金額"].values)
    select_data_qnt = sum(select_data["数量"].values)    
    
    Unit_Price_Data = pd.DataFrame({"店舗名":[select_data_shop],"数量":[select_data_qnt],"金額":[select_data_UP]})
    Unit_Rank_list.append(Unit_Price_Data)
  
  Concat_Unit_Rank_list = pd.concat(Unit_Rank_list).sort_values("金額",ascending=False).head(20)

  counter = 0
  last_row = ws_UP2.max_row
  print(last_row)
  y = str(filter_data_all["営業日付"].values[0])[0:4]
  m = str(filter_data_all["営業日付"].values[0])[4:6]
  d = str(filter_data_all["営業日付"].values[0])[6:]
  
  ymd = str(int(y)) + "/" + str(int(m)) + "/" + str(int(d))
  print(ymd)
  for row_n in Concat_Unit_Rank_list.values:
    ws_UP2["A" + str(last_row + 1 + counter ) ].value = ymd
    ws_UP2["B" + str(last_row + 1 + counter ) ].value = row_n[0]
    ws_UP2["C" + str(last_row + 1 + counter ) ].value = row_n[1]
    ws_UP2["D" + str(last_row + 1 + counter ) ].value = row_n[2]
    
    counter += 1
  #保存1
  wb_UP.save(OUT_FILE_PATH)  
  
  #ランキング順位調整
  OutWb = pyxl.load_workbook(OUT_FILE_PATH)
  OutWs = OutWb[str(int(m)) + "月"]
  r_file = pd.read_excel(OUT_FILE_PATH,sheet_name="DataBase")
  df_r_file = pd.DataFrame(r_file)
  
  df_r_file["日付"] = pd.to_datetime(df_r_file["日付"])
  
  #★
  df_r_file_filter = df_r_file[df_r_file["日付"] > yestaday]
  
  
  df_r_file_filter["日付"] = df_r_file_filter["日付"].astype(str)
  sort_df_r_file = df_r_file_filter.sort_values("金額",ascending=False)
  print("ランキング調整",sort_df_r_file)
  
  
  #TOKEN = 'NxPDQg0tpI4oY6BZHo7vkZ3gxPtTIijpWFyN85xL2q1'#テストの部屋トークン
  TOKEN = 'TNKXBcpEMmK4JAmRPaOyVABkA5GWIJkIHQOnsyfu4MD'#FUNの部屋トークン
  api_url = 'https://notify-api.line.me/api/notify'
  headers = {'Authorization' : 'Bearer ' + TOKEN}
  
  #記録更新リアクション
  actions = []
  
  for no in range(20):
    print(str(sort_df_r_file["日付"].values[no]))
    #str(sort_df_r_file["日付"].values[no])
    
    OutWs["H" + str(4 + no ) ].value = str(sort_df_r_file["日付"].values[no])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[no])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[no])[8:10]))
    OutWs["E" + str(4 + no ) ].value = sort_df_r_file["店舗"].values[no]
    OutWs["F" + str(4 + no ) ].value = sort_df_r_file["点数"].values[no]
    OutWs["G" + str(4 + no ) ].value = sort_df_r_file["金額"].values[no]
  
    if  str(str(sort_df_r_file["日付"].values[no])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[no])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[no])[8:10]))) == str(ToDay1) :
    #if  str(sort_df_r_file["日付"].values[no]) == str(ToDay1) :
      act = "🌟 NEW RANKIN 🌟"
      
    else :
      act = ""
      
    actions.append(act)  
  print(actions)  
  #保存2
  OutWb.save(OUT_FILE_PATH)
  
  message_1 = ('\n'+ str(int(m)) + '月 / 月間客単価ランキング BEST10'+'\n'+
               ToDay2 + "時点暫定結果" + "\n"
              "1位〜10位"+'\n'+'\n'+
            
            '👑1位  {}'.format(actions[0])  + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[0]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[0])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[0]) + '\n' + 
            ' 買上日:  ' + str(sort_df_r_file["日付"].values[0])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[0])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[0])[8:10])) + '\n\n' + 
            
            '👑2位  {}'.format(actions[1]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[1]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[1])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[1]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[1])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[1])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[1])[8:10])) + '\n\n' + 
            
            '👑3位  {}'.format(actions[2]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[2]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[2])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[2]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[2])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[2])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[2])[8:10])) + '\n\n' + 
            
            '  4位  {}'.format(actions[3]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[3]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[3])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[3]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[3])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[3])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[3])[8:10])) + '\n\n' + 
            
            '  5位  {}'.format(actions[4]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[4]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[4])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[4]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[4])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[4])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[4])[8:10])) + '\n\n' + 
            
            '  6位  {}'.format(actions[5]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[5]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[5])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[5]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[5])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[5])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[5])[8:10])) + '\n\n' + 
            
            '  7位  {}'.format(actions[6]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[6]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[6])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[6]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[6])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[6])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[6])[8:10])) + '\n\n' + 
            
            '  8位  {}'.format(actions[7]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[7]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[7])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[7]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[7])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[7])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[7])[8:10])) + '\n\n' + 
            
            '  9位  {}'.format(actions[8]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[8]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[8])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[8]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[8])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[8])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[8])[8:10])) + '\n\n' + 
            
            '  10位  {}'.format(actions[9]) + '\n' +
            '販売金額:  ¥' + str('{: ,}'.format(int(sort_df_r_file["金額"].values[9]))) + '\n' + 
            '販売点数:  '+ str(int(sort_df_r_file["点数"].values[9])) + '点\n' +
            ' 店舗:  ' + str(sort_df_r_file["店舗"].values[9]) + '\n' + 
            ' 買上日:  '+ str(sort_df_r_file["日付"].values[9])[0:4] + "/" + str(int(str(sort_df_r_file["日付"].values[9])[5:7])) + "/" + str(int(str(sort_df_r_file["日付"].values[9])[8:10])) + '\n\n' + 
            
        "\n"    
  )
  
  payload = {'message': message_1}
  requests.post(api_url, headers=headers, params=payload)   

    

#アイテム別平均単価
Item_Price_AVG = {}

for cd_id in division_CD :
  key_data_n =filter_data_all[filter_data_all["アイテムCD"] == cd_id]
  med = np.median(key_data_n["金額"].values)
  #filter_1 = key_data_n[key_data_n["金額"] >= 1000]

  try:
    std= int(np.std(key_data_n['金額'].values))
    
  except ValueError:
    std = 0  
  try:
      
      
    filter_ = int(np.average(key_data_n["金額"].values))
  except ValueError:
    filter_ = 0
  
  #print(cd_id,'中央値',med,"平均値",filter_,'標準偏差',std)
  print("価格チェック",filter_,cd_id)
  
  if item_underprice_list[cd_id] >= filter_ :
      filter_ = item_underprice_list[cd_id]
      
  else :
      filter_ = filter_    
      
  Item_Price_AVG[cd_id] = filter_
  
print("平均単価\n",Item_Price_AVG)  
int(np.average(filter_data_all["金額"].values))

#用途別カテゴリー平均単価を設定
op_set_price = ((Item_Price_AVG["01"] + Item_Price_AVG["12"])/2)
tops_price = ((Item_Price_AVG["04"] + Item_Price_AVG["05"] + Item_Price_AVG["07"]+ Item_Price_AVG["10"])/4)
bottoms_price = ((Item_Price_AVG["08"] + Item_Price_AVG["09"])/2)
outer_price = ((Item_Price_AVG["02"] + Item_Price_AVG["03"]+ Item_Price_AVG["06"])/3)
inn_price = 1599#Item_Price_AVG["11"] #固定値
acc_price = 799#Item_Price_AVG["13"]　#固定値
sh_price = Item_Price_AVG["15"]

print("OP/SET⇒",op_set_price,"\nTOPs⇒",tops_price,"\nBOTTTOMs⇒",bottoms_price,"\nOUTER⇒",outer_price,"\nINNER→",inn_price,"\nACC⇒",acc_price,"\nSH⇒",sh_price)



def Price_Analytics():
  Price_zone = []
  
  #pat1 = ～999 S+
  #pat2 = 1299～1999 S 
  #pat3 = 2299～2999 P
  #pat4 = 3299～3999 P+
  #pat5 = 3299～3999 P++
  
  #1顧客の平均単価と標準偏差でクラスわけし、ターゲットゾーンを選定

    
  

  # for p in Price_list:
  #   print(p)
  #   quantity = len(filter_data[filter_data["金額"] == p])
    
  #   p_row = pd.DataFrame({"価格": [p],"販売点数": [quantity]})
  #   print(p_row)
  #   Price_zone.append(p_row)
    
    
  # concat_price_zone = pd.concat(Price_zone)  
  # print(concat_price_zone)  
  
def Cust_Analytics():
  cust_type = []

  for c_cd in CustCD_list:

    Basket = filter_data[filter_data["伝票番号"] == c_cd ]
    print(Basket)
    #価格
    
    #コーディネートパターン
    
    #
SET_PATTURN_1_LIST_ALL = [] #OP/SET` + ACC`
SET_PATTURN_2_LIST_ALL = [] #OP/SET` + OUTER`
SET_PATTURN_3_LIST_ALL = [] #TOPS` + BOTTOMS`
SET_PATTURN_4_LIST_ALL = [] #TOPS` + INN`
SET_PATTURN_5_LIST_ALL = [] #TOPS` + `OUTER
SET_PATTURN_6_LIST_ALL = [] #OUTER` + BOTTOMS`
SET_PATTURN_7_LIST_ALL = [] #ACC` + BOTTOMS`
SET_PATTURN_8_LIST_ALL = [] #OP/SET` + ACC + OUTER
SET_PATTURN_9_LIST_ALL = [] #TOPS` + `OUTER + OUTER + (ACC or INN)
SET_PATTURN_10_LIST_ALL = []    

S_P_LIST = []#S/P対比構成比を格納
def FD(file,Shop):# 度数分布 【 Frequency Distributiion 】
  UNQ_CD = np.unique(file["伝票番号"].values)
  
  
  S_P = []
  C_TYPE = []#Cust_Type
  for cd_key in UNQ_CD:
    basket = []
    select_order = file[file["伝票番号"] == cd_key ]
    

    for item_n in select_order.values:
      # print("要素１",int(item_n[5]))
      # print("要素２",int(item_underprice_list[item_n[3]]))
      if int(item_n[5]) >= int(item_underprice_list[item_n[3]]):
        basket.append("P")
      elif int(item_n[5]) < int(item_underprice_list[item_n[3]]):
        basket.extend("S")
    count = len(basket)
  
    try:
      S_count = sum(x == "S" for x in basket)
    except ValueError:
      S_count = 0
        

    try :    
      S_ratio = S_count / count   
    except :
      S_ratio = 0
      
    S_P.append(S_ratio)   
    
    
    
  print(S_P)
  
  for r_x in S_P:
    if r_x > 0.8:
      TYPE = "E"
    elif r_x > 0.6:
      TYPE = "D"
      
    elif r_x > 0.4:
      TYPE = "C"  
      
    elif r_x > 0.2:
      TYPE = "B"  
        
    else :
      TYPE = "A"  
      
    C_TYPE.append(TYPE)  

  
  
  C_Type_List = pd.DataFrame({"A":[sum(x == "A" for x in C_TYPE)],"B":[sum(x == "B" for x in C_TYPE)],"C":[sum(x == "C" for x in C_TYPE)],"D":[sum(x == "D" for x in C_TYPE)],"E":[sum(x == "E" for x in C_TYPE)]})
  print(C_Type_List)
  print(sum(C_Type_List.values[0]))
  
  
  A_List = ([sum(x == "A" for x in C_TYPE)])
  print(A_List)
  A_List = "{: .1f}".format(([sum(x == "A" for x in C_TYPE)][0]/sum(C_Type_List.values[0]))*100)
  B_List = "{: .1f}".format(([sum(x == "B" for x in C_TYPE)][0]/sum(C_Type_List.values[0]))*100)

                            
  C_List = "{: .1f}".format(([sum(x == "C" for x in C_TYPE)][0]/sum(C_Type_List.values[0]))*100)
  D_List = "{: .1f}".format(([sum(x == "D" for x in C_TYPE)][0]/sum(C_Type_List.values[0]))*100)
  E_List = "{: .1f}".format(([sum(x == "E" for x in C_TYPE)][0]/sum(C_Type_List.values[0]))*100)
  
  C_Type_List_Ratio = pd.DataFrame({"店舗":[Shop[1]],"A":[A_List],"B":[B_List],"C":[C_List],"D":[D_List],"E":[E_List]})
  print(C_Type_List_Ratio)
  S_P_LIST.append(C_Type_List_Ratio)
      
  
  print(C_Type_List)
    
      
        
        
        
      
      

  
      
for shop_key in tenpo:
  print(shop_key)
  
  #customer_data = pd.read_csv('C:/Users/fun-f/Desktop/analysis/data_folder/' + shop_key[1] + '顧客データ.csv',encoding='cp932')#今週実績 
  
  file_path1 = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/myfile/dataf2/" + shop_key[1] + "販売伝票明細.csv"#販売分析ログ

  r_file1 = pd.read_csv(file_path1,encoding="cp932")
  df_r_file1 = pd.DataFrame(r_file1)


  order_n = pd.DataFrame(df_r_file1["伝票番号"],columns=["伝票番号"])
  item_cd = pd.DataFrame(df_r_file1["商品コード"].astype('str').str.zfill(10).str[:10].values,columns=["商品CD"])
  item_name = pd.DataFrame(df_r_file1["商品名"],columns=["商品名"])
  category_cd = pd.DataFrame(df_r_file1["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
  quantity = pd.DataFrame(df_r_file1["数量"].values,columns=["数量"])
  amount = pd.DataFrame(df_r_file1["小計金額"].values,columns=["金額"])
  cost = pd.DataFrame(df_r_file1["原価"].values,columns=["原価"])#原価

  set_data = pd.concat([order_n,item_cd,item_name,category_cd,quantity,amount,cost],axis=1)
    
  #filter_data = set_data[set_data["金額"] >= 100]

  filter_1 = set_data[set_data["アイテムCD"] != "98"] #ショッパー除外

  filter_2 = filter_1[filter_1["アイテムCD"] != "14"] #サンプル除外

  #追加コード
  filter_3 = filter_2[filter_2["金額"] > 0 ]#返品データを削除

  filter_data = filter_3[filter_3["金額"] != 50] #マスク除外

  #差し引きデータ
  Pull_data = filter_2[filter_2["金額"] < 0 ]
  
  FD(filter_data,shop_key)
  
  #concat_S_P_LIST = pd.concat(S_P_LIST)
  


  CustCD_list = np.unique(filter_data["伝票番号"].values)
  Price_list = np.unique(filter_data["金額"].values)    


  def CustPriceZone_Analytics():

    #アイテム別平均単価
    Item_Price_AVG = {}
    CUST_TYPE = []

    for cd in CustCD_list :
      key_data_n =filter_data[filter_data["伝票番号"] == cd]
      qnt = len(key_data_n["金額"].values)#点数
      apper = max(key_data_n["金額"].values)#最大金額
      under = min(key_data_n["金額"].values)#最小金額
      avg = np.average(key_data_n["金額"].values)#平均単価
      std = np.std(key_data_n["金額"].values)#標準偏差
      
      #print("点数",qnt,"\n上限",apper,"下限",under,"\n平均単価",avg,"\n標準偏差",std)
      
      #filter_1 = key_data_n[key_data_n["金額"] >= 1000]
      try:
        std= int(np.std(key_data_n['金額'].values))
        
      except ValueError:
        std = 0  
      #try:
    #     filter_ = int(np.average(key_data_n["金額"].values))
    #   except ValueError:
    #     filter_ = 0
      
    #   #print(cd_id,'中央値',med,"平均値",filter_,'標準偏差',std)
    #   Item_Price_AVG[cd_id] = filter_
      
    # print("平均単価\n",Item_Price_AVG)  
    # int(np.average(filter_data["金額"].values))

    # op_set_price = ((Item_Price_AVG["01"] + Item_Price_AVG["12"])/2)
    # tops_price = ((Item_Price_AVG["04"] + Item_Price_AVG["05"] + Item_Price_AVG["07"]+ Item_Price_AVG["10"])/4)
    # bottoms_price = ((Item_Price_AVG["08"] + Item_Price_AVG["09"])/2)
    # outer_price = ((Item_Price_AVG["02"] + Item_Price_AVG["03"]+ Item_Price_AVG["06"])/3)
    # inn_price = Item_Price_AVG["11"]
    # acc_price = Item_Price_AVG["13"]
    # sh_price = Item_Price_AVG["15"]
  CustPriceZone_Analytics()   


  #顧客データを格納
  CUST_DATA_LIST = {}
  CUST_TYPE_1 = 0#相関関係あり
  CUST_TYPE_2= 0#相関関係なし　
  CUST_TYPE_3= 0#単品購入客数
  CUST_TYPE_4 = 0

  SET_PATTURN_1 = 0 #OP/SET` + ACC`
  SET_PATTURN_2 = 0 #OP/SET` + OUTER`
  SET_PATTURN_3 = 0 #TOPS` + BOTTOMS`
  SET_PATTURN_4 = 0 #TOPS` + INN`
  SET_PATTURN_5 = 0 #TOPS` + `OUTER
  SET_PATTURN_6 = 0 #OUTER` + BOTTOMS`
  SET_PATTURN_7 = 0 #ACC` + BOTTOMS`
  SET_PATTURN_8 = 0 #OP/SET` + ACC + OUTER
  SET_PATTURN_9 = 0 #TOPS` + `OUTER + OUTER + (ACC or INN)
  SET_PATTURN_10 = 0

  SET_PATTURN_1_LIST = [] #OP/SET` + ACC`
  SET_PATTURN_2_LIST = [] #OP/SET` + OUTER`
  SET_PATTURN_3_LIST = [] #TOPS` + BOTTOMS`
  SET_PATTURN_4_LIST = [] #TOPS` + INN`
  SET_PATTURN_5_LIST = [] #TOPS` + `OUTER
  SET_PATTURN_6_LIST = [] #OUTER` + BOTTOMS`
  SET_PATTURN_7_LIST = [] #ACC` + BOTTOMS`
  SET_PATTURN_8_LIST = [] #OP/SET` + ACC + OUTER
  SET_PATTURN_9_LIST = [] #TOPS` + `OUTER + OUTER + (ACC or INN)
  SET_PATTURN_10_LIST = []

  for i in CustCD_list:
    key_data = filter_data[filter_data["伝票番号"] == i]#伝票番番号に一致するデータを抽出
    data_at = sum(key_data["金額"].values)#客単価
    data_count = len(key_data)#買上数をカウント
    #基準値
    #P_ALL ⇒ A
    #P:S 5:5 ⇒ B
    #
    
    
    op_set_list = []
    tops_list = []
    bottoms_list = []
    outer_list = []
    inn_list = []
    acc_list = []
    sh_list = []
    
    
    #提案の相関関係が津陽
    #タイプA
    
    P = 0 
    S = 0
    #一致伝票のデータから該当アイテムを取得
    for i_2,n in zip(key_data.values,range(1,data_count + 1 )):
      
      if (i_2[6]/i_2[5]) < 0.3 :
        S += 1
        
      else:
        P += 1  
        
      
      if (i_2[3] == "01") & (i_2[5] >= Item_Price_AVG["01"]) or (i_2[3] == "12") & (i_2[5] >= Item_Price_AVG["12"]):
        op_set_list.append(i_2)
      
      elif (i_2[3] == "04") & (i_2[5] >= Item_Price_AVG["04"]) or (i_2[3] == "05") & (i_2[5] >= Item_Price_AVG["05"]) or (i_2[3] == "07") & (i_2[5] >= Item_Price_AVG["07"]) or (i_2[3] == "10") & (i_2[5] >= Item_Price_AVG["10"]):
        
        tops_list.append(i_2)
      
      elif ((i_2[3] == "08") & (i_2[5] >= Item_Price_AVG["08"])) or ((i_2[3] == "09") & (i_2[5] >= Item_Price_AVG["09"])) :
        
        bottoms_list.append(i_2)  
        
      elif (i_2[3] == "02") & (i_2[5] >= Item_Price_AVG["02"]) or (i_2[3] == "03") & (i_2[5] >= Item_Price_AVG["03"]) or (i_2[3] == "06") & (i_2[5] >= Item_Price_AVG["06"]): 
        
        outer_list.append(i_2)
        
      elif (i_2[3] == "11") & (i_2[5] >= Item_Price_AVG["11"]):
        
        inn_list.append(i_2)
      
      elif (i_2[3] == "13") & (i_2[5] >= Item_Price_AVG["13"]):
        
        acc_list.append(i_2)      
      
      elif (i_2[3] == "15") & (i_2[5] >= Item_Price_AVG["15"]):
        
        sh_list.append(i_2)
        
    #提案の相関関係が津陽
    #タイプA
    #2点SETの組み合わせ
    

    if data_count >= 2:
      
      if (data_count >= 4) & (data_at >= 7500):
        type_n = "S"  
        CUST_TYPE_4 += 1
        if (len(op_set_list) > 0 ) & (len(acc_list) > 0) & (len(outer_list) > 0) :
          SET_PATTURN_8 += 1
          SET_PATTURN_8_LIST.append(key_data)
        elif (len(tops_list) > 0 ) & (len(bottoms_list) > 0) & (len(outer_list)) & ((len(acc_list)) or (len(inn_list))):
          SET_PATTURN_9 += 1
          SET_PATTURN_9_LIST.append(key_data)
        
        
        
      #OP提案の判定
      elif (len(op_set_list) > 0 ) & (len(acc_list) > 0) & (data_at >= (op_set_price + acc_price)):
        type_n = "A"
        CUST_TYPE_1 += 1
        SET_PATTURN_1 += 1
        SET_PATTURN_1_LIST.append(key_data)
        
      elif (len(op_set_list) > 0 ) & (len(outer_list) > 0) & (data_at >= (op_set_price + outer_price)):
        type_n = "A"
        CUST_TYPE_1 += 1
        SET_PATTURN_2 += 1
        SET_PATTURN_2_LIST.append(key_data)
        
      #Tops + Bottoms 提案を判定
      elif (len(tops_list) > 0 ) & (len(bottoms_list) > 0) & (data_at >= (tops_price + bottoms_price)) :
        type_n = "A"  
        CUST_TYPE_1 += 1
        SET_PATTURN_3 += 1
        SET_PATTURN_3_LIST.append(key_data)
        
      #Tops + INN 提案を判定
      elif (len(tops_list) > 0 ) & (len(inn_list) > 0) & (data_at >= (tops_price + inn_price)) :
        type_n = "A"    
        CUST_TYPE_1 += 1
        SET_PATTURN_4 += 1
        SET_PATTURN_4_LIST.append(key_data)
      #Outer + Tops 提案を判定
      elif (len(outer_list) > 0 ) & (len(tops_list) > 0) & (data_at >= (outer_price + tops_price)):
        type_n = "A"   
        CUST_TYPE_1 += 1
        SET_PATTURN_5 += 1
        SET_PATTURN_5_LIST.append(key_data)
        
      #Outer + Tops 提案を判定
      elif (len(outer_list) > 0 ) & (len(bottoms_list) > 0) & (data_at >= (outer_price + bottoms_price)):
        type_n = "A"   
        CUST_TYPE_1 += 1
        SET_PATTURN_6 += 1
        SET_PATTURN_6_LIST.append(key_data)
      
      #Bottoms + Acc 提案を判定
      elif (len(bottoms_list) > 0 ) & (len(acc_list) > 0) & (data_at >= (bottoms_price + acc_price)):
        type_n = "A"   
        CUST_TYPE_1 += 1  
        SET_PATTURN_7 += 1
        SET_PATTURN_7_LIST.append(key_data)
        
      else :
        type_n = "相関関係不明"  
        CUST_TYPE_2 += 1
    
      try:  
        PSR = "{: .1f}".format((P/data_count)*100)
      except ZeroDivisionError:
        PSR = 0  
      #cust_data = {i:[[type_n,data_count,data_at],[op_set_list,tops_list,bottoms_list,outer_list,inn_list,acc_list,sh_list]]} 
      CUST_DATA_LIST[str(i)] = [[type_n,data_count,data_at,PSR],[op_set_list,tops_list,bottoms_list,outer_list,inn_list,acc_list,sh_list]]
    
    else :
      type_n = "相関関係無し"  
      CUST_TYPE_3 += 1
      
      try:  
        PSR = "{: .1f}".format((P/data_count)*100)
      except ZeroDivisionError:
        PSR = 0  
      
      CUST_DATA_LIST[str(i)] = [[type_n,data_count,data_at,PSR],[op_set_list,tops_list,bottoms_list,outer_list,inn_list,acc_list,sh_list]]
      
  # CONCAT_SET_PATTURN_1_LIST = pd.concat(SET_PATTURN_1_LIST)
  # CONCAT_SET_PATTURN_2_LIST = pd.concat(SET_PATTURN_2_LIST)
  # CONCAT_SET_PATTURN_3_LIST = pd.concat(SET_PATTURN_3_LIST)
  # CONCAT_SET_PATTURN_4_LIST = pd.concat(SET_PATTURN_4_LIST)
  # CONCAT_SET_PATTURN_5_LIST = pd.concat(SET_PATTURN_5_LIST)
  # CONCAT_SET_PATTURN_6_LIST = pd.concat(SET_PATTURN_6_LIST)
  # CONCAT_SET_PATTURN_7_LIST = pd.concat(SET_PATTURN_7_LIST)
  # CONCAT_SET_PATTURN_8_LIST = pd.concat(SET_PATTURN_8_LIST)
  # CONCAT_SET_PATTURN_9_LIST = pd.concat(SET_PATTURN_9_LIST)


  ALL_CUST = CUST_TYPE_1 + CUST_TYPE_2 + CUST_TYPE_3 + CUST_TYPE_4

  # ws["B" + str(3 + counter_n)].value = CUST_TYPE_1 + CUST_TYPE_4
  # #ws["C" + str(3 + counter_n)].value = "{:.1f}".format(((CUST_TYPE_1 + CUST_TYPE_4)/ALL_CUST)*100)
  # ws["D" + str(3 + counter_n)].value = CUST_TYPE_2
  # #ws["E" + str(3 + counter_n)].value = "{:.1f}".format(((CUST_TYPE_2)/ALL_CUST)*100)
  # ws["F" + str(3 + counter_n)].value = CUST_TYPE_3
  # #ws["G" + str(3 + counter_n)].value = "{:.1f}".format((CUST_TYPE_3/ALL_CUST)*100)
  # ws["H" + str(3 + counter_n)].value = CUST_TYPE_4
  # #ws["I" + str(3 + counter_n)].value = "{:.1f}".format((CUST_TYPE_4/ALL_CUST)*100)

  # ws["R" + str(3 + counter_n)].value = SET_PATTURN_1
  # ws["S" + str(3 + counter_n)].value = SET_PATTURN_2
  # ws["T" + str(3 + counter_n)].value = SET_PATTURN_3
  # ws["U" + str(3 + counter_n)].value = SET_PATTURN_4
  # ws["V" + str(3 + counter_n)].value = SET_PATTURN_5
  # ws["W" + str(3 + counter_n)].value = SET_PATTURN_6
  # ws["X" + str(3 + counter_n)].value = SET_PATTURN_7
  # ws["Y" + str(3 + counter_n)].value = SET_PATTURN_8
  # ws["Z" + str(3 + counter_n)].value = SET_PATTURN_9
  

  print(
        "相関関係あり⇒",CUST_TYPE_1 + CUST_TYPE_4,"{:.1f}".format(((CUST_TYPE_1 + CUST_TYPE_4)/ALL_CUST)*100),
        "4点以上購入客数⇒",CUST_TYPE_4,"{:.1f}".format((CUST_TYPE_4/ALL_CUST)*100),
        "相関関係不明⇒",CUST_TYPE_2,"{:.1f}".format(((CUST_TYPE_2)/ALL_CUST)*100),
        "単品購入客数⇒",CUST_TYPE_3,"{:.1f}".format((CUST_TYPE_3/ALL_CUST)*100),
        "PSR⇒",S,
        )
  #ショッパー抜き顧客買上データ
  #第1項目
  #
  #第2項目
  #
  

  print("コーデパターン集計\n",
      "\nOP/SET + ACC⇒",SET_PATTURN_1,#第1項目
      "\nOP/SET + OUTER⇒",SET_PATTURN_2,
      "\nTOPS + BOTTOMS⇒",SET_PATTURN_3,#第1項目
      "\nTOPS + INN⇒",SET_PATTURN_4,
      "\nTOPS + OUTER⇒",SET_PATTURN_5,
      "\nOUTER + BOTTOMS⇒",SET_PATTURN_6,
      "\nACC + BOTTOMS⇒",SET_PATTURN_7,
      "\nOP/SET + ACC + OUTER⇒",SET_PATTURN_8,#第3項目
      "\nTOPS + BOTTOMS + OUTER + (ACC or INN)⇒",SET_PATTURN_9,
      
  )

  
  print(len(SET_PATTURN_1_LIST),len(SET_PATTURN_2_LIST),len(SET_PATTURN_3_LIST),len(SET_PATTURN_4_LIST),len(SET_PATTURN_5_LIST),len(SET_PATTURN_6_LIST),len(SET_PATTURN_7_LIST),)
  if len(SET_PATTURN_1_LIST) > 0:
    SET_PATTURN_1_LIST_ALL.append(pd.concat(SET_PATTURN_1_LIST))
    
  if len(SET_PATTURN_2_LIST) > 0:
    SET_PATTURN_2_LIST_ALL.append(pd.concat(SET_PATTURN_2_LIST))
    
  if len(SET_PATTURN_3_LIST) > 0:
    SET_PATTURN_3_LIST_ALL.append(pd.concat(SET_PATTURN_3_LIST))
    
  if len(SET_PATTURN_4_LIST) > 0:
    SET_PATTURN_4_LIST_ALL.append(pd.concat(SET_PATTURN_4_LIST))
    
  if len(SET_PATTURN_5_LIST) > 0:
    SET_PATTURN_5_LIST_ALL.append(pd.concat(SET_PATTURN_5_LIST))
    
  if len(SET_PATTURN_6_LIST) > 0:
    SET_PATTURN_6_LIST_ALL.append(pd.concat(SET_PATTURN_6_LIST))
    
  if len(SET_PATTURN_7_LIST) > 0:
    SET_PATTURN_7_LIST_ALL.append(pd.concat(SET_PATTURN_7_LIST))          
#★★    


# CONCAT_SET_PATTURN_1_LIST_ALL = pd.concat(SET_PATTURN_1_LIST_ALL)
# SET_PATTURN_1CNT = len(np.unique(CONCAT_SET_PATTURN_1_LIST_ALL["伝票番号"].values))
# SET_PATTURN_1VAL = sum(CONCAT_SET_PATTURN_1_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_1VAL)
# print("SETパターン1",SET_PATTURN_1CNT)
# # for set_cd in SET_PATTURN_1CNT:
# #   select_row = CONCAT_SET_PATTURN_1_LIST_ALL[CONCAT_SET_PATTURN_1_LIST_ALL["伝票番号"] == set_cd]


# CONCAT_SET_PATTURN_2_LIST_ALL = pd.concat(SET_PATTURN_2_LIST_ALL)
# SET_PATTURN_2CNT = len(np.unique(CONCAT_SET_PATTURN_2_LIST_ALL["伝票番号"].values))
# SET_PATTURN_2VAL = sum(CONCAT_SET_PATTURN_2_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_2VAL)
# print("SETパターン2",SET_PATTURN_2CNT)


# CONCAT_SET_PATTURN_3_LIST_ALL = pd.concat(SET_PATTURN_3_LIST_ALL)
# SET_PATTURN_3CNT = len(np.unique(CONCAT_SET_PATTURN_3_LIST_ALL["伝票番号"].values))
# SET_PATTURN_3VAL = sum(CONCAT_SET_PATTURN_3_LIST_ALL["金額"].values)

# #for item in CONCAT_SET_PATTURN_3_LIST_ALL.values:
#   #print(item)
# print("金額",SET_PATTURN_3VAL)
# print("SETパターン3",SET_PATTURN_3CNT)

# CONCAT_SET_PATTURN_4_LIST_ALL = pd.concat(SET_PATTURN_4_LIST_ALL)
# SET_PATTURN_4CNT = len(np.unique(CONCAT_SET_PATTURN_4_LIST_ALL["伝票番号"].values))
# SET_PATTURN_4VAL = sum(CONCAT_SET_PATTURN_4_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_4VAL)
# print("SETパターン4",SET_PATTURN_4CNT)

# CONCAT_SET_PATTURN_5_LIST_ALL = pd.concat(SET_PATTURN_5_LIST_ALL)
# SET_PATTURN_5CNT = len(np.unique(CONCAT_SET_PATTURN_5_LIST_ALL["伝票番号"].values))
# SET_PATTURN_5VAL = sum(CONCAT_SET_PATTURN_5_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_5VAL)
# print("SETパターン5",SET_PATTURN_5CNT)


# CONCAT_SET_PATTURN_6_LIST_ALL = pd.concat(SET_PATTURN_6_LIST_ALL)
# SET_PATTURN_6CNT = len(np.unique(CONCAT_SET_PATTURN_6_LIST_ALL["伝票番号"].values))
# SET_PATTURN_6VAL = sum(CONCAT_SET_PATTURN_6_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_6VAL)
# print("SETパターン6",SET_PATTURN_6CNT)


# CONCAT_SET_PATTURN_7_LIST_ALL = pd.concat(SET_PATTURN_7_LIST_ALL)
# SET_PATTURN_7CNT = len(np.unique(CONCAT_SET_PATTURN_7_LIST_ALL["伝票番号"].values))
# SET_PATTURN_7VAL = sum(CONCAT_SET_PATTURN_7_LIST_ALL["金額"].values)
# print("金額",SET_PATTURN_7VAL)
# print("SETパターン7",SET_PATTURN_7CNT)


# unq_PATTURN_3 = np.unique(CONCAT_SET_PATTURN_3_LIST_ALL["商品CD"].values)
# rank_PATTURN_3 = []
# for u_cd in unq_PATTURN_3:
#   key_data_p3 = CONCAT_SET_PATTURN_3_LIST_ALL[CONCAT_SET_PATTURN_3_LIST_ALL["商品CD"] == u_cd]

#   key_datarow_p3 = pd.DataFrame({
#                                 "商品CD":[u_cd],
#                                 "商品名":[key_data_p3["商品名"].values[0]],
#                                 "点数":[sum(key_data_p3["数量"].values)],
#                                 "金額":[sum(key_data_p3["金額"].values)],
#                                  })
  
#   rank_PATTURN_3.append(key_datarow_p3)

# concat_rank_PATTURN_3 = pd.concat(rank_PATTURN_3).sort_values("金額",ascending=False)
# print(concat_rank_PATTURN_3)  
# concat_S_P_LIST = pd.concat(S_P_LIST)
# print(concat_S_P_LIST)

#★★    
  
UnitPrice()  





  
